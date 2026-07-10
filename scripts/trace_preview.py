import sys
from pathlib import Path
import tempfile
import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from src.core.summary_reconciliation_engine import SummaryReconciliationEngine

def find_line(filename, needle):
    for i, l in enumerate(Path(filename).read_text().splitlines(), 1):
        if needle in l:
            return i, l.strip()
    return None, None

def main():
    # create temp workbook with formula cells in summary
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws_daily = wb.create_sheet('DAILY OUTPUT FILE 03-07-2026')
    ws_daily.append(['Creation_Date', 'Branch_Code', 'Branch_Name', 'Card_Type', 'No_of_Batches', 'Waybill_No', 'Batch_No', 'Bag_No'])
    ws_daily.append(['2026-07-03', 'BR003', 'Branch C', 'SIM', 1, 'WB003', '30001', "'00034567"])

    ws = wb.create_sheet('CAPITEC SUMMARY FILE REPORT')
    ws.append(['Files received', 'Total quantity received', 'Quantity in stock', 'Quantity dispatched', 'Comment'])
    # Put formulas in stock/dispatched to simulate real workbook
    ws.append(['C-Connect batch 1', 800, '=500', '=100', 'Active SIM record'])

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()
    path = tmp.name
    wb.save(path)

    engine = SummaryReconciliationEngine()
    engine.loader.load_workbook(path)
    summary_sheet = engine.loader.detect_capitec_summary_sheet()
    worksheet = engine.loader.get_worksheet(summary_sheet)

    # locate header row
    header_row_idx, headers = engine._find_summary_header_row(worksheet)
    normalized_header_map = {engine._normalize_header(h): idx+1 for idx,h in enumerate(headers) if h}
    files_col = normalized_header_map.get('files received',1)
    stock_col = normalized_header_map.get('quantity in stock')
    dispatched_col = normalized_header_map.get('quantity dispatched')

    # iterate rows to reproduce _read_summary_rows behavior and trace
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=header_row_idx+1), start=header_row_idx+1):
        item_cell = row[files_col-1] if len(row) >= files_col else None
        item_name = str(item_cell.value).strip() if item_cell and item_cell.value is not None else ''
        if not item_name:
            continue

        # Step 1: raw cell value read
        raw_stock = engine._get_raw_cell_value(row, stock_col)
        ln, code = find_line('src/core/summary_reconciliation_engine.py', 'def _get_raw_cell_value')
        print(f"File: src/core/summary_reconciliation_engine.py | Function: _get_raw_cell_value | Variable: raw_quantity_in_stock | Line: {ln} | Value: {raw_stock}")

        # Step 2: SummaryRow.quantity_in_stock parsed
        qty_stock = engine._parse_int_cell(row, stock_col)
        ln2, code2 = find_line('src/core/summary_reconciliation_engine.py', 'def _parse_int_cell')
        print(f"File: src/core/summary_reconciliation_engine.py | Function: _parse_int_cell | Variable: quantity_in_stock | Line: {ln2} | Value: {qty_stock}")

        if qty_stock is None and raw_stock is not None:
            reason = 'Cell contains an Excel formula (leading =), parser does not handle formulas; conversion failed'
            print(f"STOP: File: src/core/summary_reconciliation_engine.py | Function: _parse_int_cell | Variable: quantity_in_stock | Line: {ln2} | Why it became None: {reason}")
            return

    # If not stopped, also show SummaryAnalysis storage and dashboard usage
    analysis = engine.analyze(path)
    # Step 3: SummaryAnalysis stores quantity_in_stock
    sr = analysis.summary_rows[0]
    ln3, _ = find_line('src/models/models.py', 'class SummaryRow')
    print(f"File: src/models/models.py | Function: SummaryRow.__init__ | Variable: quantity_in_stock | Line: {ln3} | Value: {sr.quantity_in_stock}")

    # Step 4: Dashboard receives analysis (line where self.summary_analysis assigned)
    ln4, _ = find_line('src/gui/dashboard.py', 'self.summary_analysis = analysis')
    print(f"File: src/gui/dashboard.py | Function: _run_integration_test | Variable: summary_analysis | Line: {ln4} | Value: {analysis.summary_rows[0].quantity_in_stock}")

    # Step 5: build_summary_update_preview_text uses row.quantity_in_stock in _build_summary_row_preview_rows
    ln5, _ = find_line('src/gui/dashboard.py', 'def _build_summary_row_preview_rows')
    print(f"File: src/gui/dashboard.py | Function: _build_summary_row_preview_rows | Variable: row.quantity_in_stock | Line: {ln5} | Value: {analysis.summary_rows[0].quantity_in_stock}")

if __name__ == '__main__':
    main()
