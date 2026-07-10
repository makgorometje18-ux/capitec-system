"""
Auto-Fix Module - Automatically fixes common Excel workbook issues.
"""

from typing import Dict, Tuple, List
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

from src.utils.logger import get_logger


class AutoFixer:
    """
    Automatically fixes common validation issues in Excel workbooks.
    
    Currently supports:
    - Auto-fix extra leading apostrophes in Bag_No values
    """
    
    def __init__(self):
        """Initialize the Auto Fixer."""
        self.logger = get_logger()
        self.workbook = None
        self.current_file_path = None
        self.changes_made = 0
    
    def load_workbook(self, file_path: str) -> bool:
        """
        Load a workbook for auto-fixing.
        
        Args:
            file_path: Path to the Excel workbook.
            
        Returns:
            True if successful, False otherwise.
        """
        try:
            path = Path(file_path)
            if not path.exists():
                self.logger.error(f"File not found: {file_path}")
                return False
            
            self.workbook = openpyxl.load_workbook(file_path)
            self.current_file_path = file_path
            self.changes_made = 0
            self.logger.info(f"Workbook loaded for auto-fixing: {file_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error loading workbook: {e}")
            return False
    
    def fix_bag_numbers(self, sheet_name: str) -> Tuple[bool, int]:
        """
        Auto-fix bag numbers with extra leading apostrophes.
        
        Business Rule BR007: Bag_No should contain exactly one leading apostrophe.
        This method fixes values with multiple apostrophes (e.g., ''00004567 → '00004567).
        
        Args:
            sheet_name: Name of the worksheet to fix.
            
        Returns:
            Tuple of (success: bool, changes_made: int)
        """
        if not self.workbook:
            self.logger.error("No workbook loaded")
            return False, 0
        
        try:
            worksheet = self.workbook[sheet_name]
            fixes = 0
            
            # Find Bag_No column header
            bag_no_col = None
            headers = []
            
            for col_idx, cell in enumerate(worksheet[1], 1):
                if cell.value:
                    headers.append((col_idx, str(cell.value).strip()))
            
            for col_idx, header in headers:
                if header == "Bag_No":
                    bag_no_col = col_idx
                    break
            
            if not bag_no_col:
                self.logger.warning(f"Bag_No column not found in {sheet_name}")
                return True, 0
            
            # Fix values in Bag_No column
            for row_idx in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=bag_no_col)
                if cell.value:
                    value = str(cell.value)
                    fixed_value = self._fix_bag_number(value)
                    
                    if fixed_value != value:
                        cell.value = fixed_value
                        fixes += 1
                        self.logger.debug(f"Fixed Bag_No at row {row_idx}: '{value}' → '{fixed_value}'")
            
            self.changes_made += fixes
            if fixes > 0:
                self.logger.info(f"Fixed {fixes} Bag_No values in {sheet_name}")
            
            return True, fixes
            
        except Exception as e:
            self.logger.error(f"Error fixing bag numbers: {e}")
            return False, 0
    
    def _fix_bag_number(self, value: str) -> str:
        """
        Fix a single bag number value.
        
        Removes extra leading apostrophes while preserving exactly one.
        
        Args:
            value: Original bag number value.
            
        Returns:
            Fixed bag number value.
        """
        if not value:
            return value
        
        # Count leading apostrophes
        leading_apostrophes = 0
        for char in value:
            if char == "'":
                leading_apostrophes += 1
            else:
                break
        
        # If more than one leading apostrophe, fix it
        if leading_apostrophes > 1:
            # Remove all leading apostrophes and add exactly one
            cleaned = value.lstrip("'")
            return f"'{cleaned}"
        
        return value
    
    def get_changes_summary(self) -> Dict[str, int]:
        """
        Get a summary of all changes made.
        
        Returns:
            Dictionary with change counts.
        """
        return {
            "total_changes": self.changes_made
        }
    
    def save_workbook(self, output_path: str = None) -> bool:
        """
        Save the workbook with fixes applied.
        
        Args:
            output_path: Path to save the workbook. If None, overwrites original.
            
        Returns:
            True if successful, False otherwise.
        """
        if not self.workbook:
            self.logger.error("No workbook loaded")
            return False
        
        try:
            save_path = output_path or self.current_file_path
            self.workbook.save(save_path)
            self.logger.info(f"Workbook saved with {self.changes_made} auto-fixes: {save_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error saving workbook: {e}")
            return False
    
    def close(self):
        """Close the workbook without saving."""
        if self.workbook:
            self.workbook.close()
            self.logger.debug("Workbook closed")
