from typing import Any, Dict, List, Optional

from openpyxl import Workbook

from src.models.models import CommercialBatch


COMMERCIAL_BATCH_SHEET_PREFIX = 'COMMERCIAL BATCH'


def _normalize_header(value: Optional[Any]) -> str:
    if value is None:
        return ''
    return str(value).strip().lower().replace(' ', '_')


def _find_commercial_batch_sheet(workbook: Workbook) -> Optional[str]:
    if workbook is None or not getattr(workbook, 'sheetnames', None):
        return None

    for sheet_name in workbook.sheetnames:
        if str(sheet_name).strip().lower().startswith(COMMERCIAL_BATCH_SHEET_PREFIX.lower()):
            return sheet_name
    return None


def _parse_int(value: Optional[Any]) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    try:
        text = str(value).strip()
        if text == '':
            return None
        text = text.replace(',', '')
        return int(float(text)) if '.' in text else int(text)
    except Exception:
        return None


def _find_header_row(worksheet) -> tuple[int, List[str]]:
    for row_idx, row in enumerate(worksheet.iter_rows(max_row=20, values_only=True), start=1):
        normalized = [_normalize_header(cell) for cell in row if cell is not None]
        if any(name in normalized for name in ('batch_number', 'batch_number', 'batchno', 'batch')) and \
           any(name in normalized for name in ('card_type', 'cardtype')) and \
           any(name in normalized for name in ('quantity', 'qty')):
            headers = [str(cell).strip() if cell is not None else '' for cell in row]
            return row_idx, headers

    first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [str(cell).strip() if cell is not None else '' for cell in first_row]
    return 1, headers


def _build_header_map(headers: List[str]) -> Dict[str, int]:
    normalized_headers = [_normalize_header(header) for header in headers]
    return {name: idx + 1 for idx, name in enumerate(normalized_headers)}


def _find_column(header_map: Dict[str, int], *aliases: str) -> Optional[int]:
    for alias in aliases:
        if alias in header_map:
            return header_map[alias]
    return None


def read_commercial_batches(workbook) -> List[CommercialBatch]:
    sheet_name = _find_commercial_batch_sheet(workbook)
    if not sheet_name:
        return []

    worksheet = workbook[sheet_name]
    header_row_idx, headers = _find_header_row(worksheet)
    header_map = _build_header_map(headers)

    batch_number_col = _find_column(header_map, 'batch_number', 'batchno', 'batch')
    card_type_col = _find_column(header_map, 'card_type', 'cardtype')
    quantity_col = _find_column(header_map, 'quantity', 'qty')
    order_number_col = _find_column(header_map, 'order_number', 'orderno', 'order_number', 'order')

    batches: List[CommercialBatch] = []
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=header_row_idx + 1), start=header_row_idx + 1):
        batch_number_cell = row[batch_number_col - 1] if batch_number_col and len(row) >= batch_number_col else None
        card_type_cell = row[card_type_col - 1] if card_type_col and len(row) >= card_type_col else None
        quantity_cell = row[quantity_col - 1] if quantity_col and len(row) >= quantity_col else None
        order_number_cell = row[order_number_col - 1] if order_number_col and len(row) >= order_number_col else None

        batch_number = str(batch_number_cell.value).strip() if batch_number_cell and batch_number_cell.value is not None else ''
        card_type = str(card_type_cell.value).strip() if card_type_cell and card_type_cell.value is not None else ''
        quantity = _parse_int(quantity_cell.value) if quantity_cell is not None else None
        order_number = str(order_number_cell.value).strip() if order_number_cell and order_number_cell.value is not None else None

        if not batch_number and not card_type and quantity is None and not order_number:
            continue

        batches.append(
            CommercialBatch(
                batch_number=batch_number,
                card_type=card_type,
                quantity=quantity if quantity is not None else 0,
                order_number=order_number,
                raw_row_index=row_idx,
            )
        )

    return batches
