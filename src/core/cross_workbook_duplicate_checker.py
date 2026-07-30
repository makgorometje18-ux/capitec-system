"""
Cross-Workbook Duplicate Checker

Scans a folder of previous Daily Output workbooks and aggregates Batch_Number
values to support cross-workbook duplicate detection.
"""
from typing import Dict, List, Tuple
from pathlib import Path
import openpyxl
from src.utils.helpers import split_batch_numbers, clean_string
from src.utils.logger import get_logger


class CrossWorkbookDuplicateChecker:
    """Checks previous workbooks for batch numbers."""

    def __init__(self, previous_folder: str):
        self.logger = get_logger()
        self.previous_folder = Path(previous_folder) if previous_folder else None
        self.previous_index: Dict[str, List[Tuple[str, str, int, str]]] = {}
        # batch -> list of (workbook_name, sheet_name, row, cell_ref)

    def index_previous_workbooks(self) -> bool:
        """Index all previous Daily Output workbooks in the configured folder.

        Returns True if indexing succeeded (even if no files), False on fatal error.
        """
        if not self.previous_folder:
            self.logger.warning("Previous workbooks folder not configured")
            return False

        if not self.previous_folder.exists():
            self.logger.warning(f"Previous workbooks folder not found: {self.previous_folder}")
            return False

        try:
            files = list(self.previous_folder.glob("*.xlsx"))
            if not files:
                self.logger.info("Previous workbooks folder is empty")
                return True

            for file in files:
                try:
                    wb = openpyxl.load_workbook(file, data_only=True)
                except Exception as e:
                    self.logger.warning(f"Failed to open previous workbook {file.name}: {e}")
                    continue

                for sheet_name in wb.sheetnames:
                    if not sheet_name.startswith("DAILY OUTPUT FILE"):
                        continue
                    ws = wb[sheet_name]
                    # find Batch_Number column index in header row
                    header_cells = list(ws[1])
                    batch_col_idx = None
                    for idx, cell in enumerate(header_cells, start=1):
                        if cell.value:
                            normalized = str(cell.value).strip().lower().replace(' ', '_')
                            if normalized == 'batch_number':
                                batch_col_idx = idx
                                break

                    if not batch_col_idx:
                        continue

                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=batch_col_idx)
                        if not cell or cell.value is None:
                            continue
                        batch_string = str(cell.value)
                        batches = split_batch_numbers(batch_string)
                        for batch in batches:
                            b = clean_string(batch)
                            if not b:
                                continue
                            entry = (file.name, sheet_name, row_idx, f"{batch_col_idx}:{row_idx}")
                            self.previous_index.setdefault(b, []).append(entry)
                wb.close()
            self.logger.info(f"Indexed previous workbooks: {len(self.previous_index)} unique batches")
            return True
        except Exception as e:
            self.logger.error(f"Error indexing previous workbooks: {e}")
            return False

    def find_previous_occurrences(self, batch_number: str) -> List[Tuple[str, str, int, str]]:
        """Return list of occurrences for a batch number from previous workbooks."""
        return self.previous_index.get(clean_string(batch_number), [])
