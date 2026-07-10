"""
Duplicate Report Generator - Creates a worksheet listing all duplicate batch numbers.
"""

from typing import List, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.models.models import DuplicateRecord
from src.utils.logger import get_logger


class DuplicateReportGenerator:
    """
    Generates a detailed Duplicate Report worksheet in the Excel workbook.
    
    Creates a new worksheet called "Duplicate Report" that lists:
    - Batch Number
    - Worksheet Name
    - Row Number
    - Cell Reference
    - Occurrences
    - Duplicate Type (Same Cell / Across Rows)
    - Description
    """
    
    REPORT_SHEET_NAME = "Duplicate Report"
    
    # Header styling
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    
    # Border styling
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self):
        """Initialize the Duplicate Report Generator."""
        self.logger = get_logger()
        self.workbook = None
        self.current_file_path = None
    
    def load_workbook(self, file_path: str) -> bool:
        """
        Load a workbook for report generation.
        
        Args:
            file_path: Path to the Excel workbook.
            
        Returns:
            True if successful, False otherwise.
        """
        try:
            from pathlib import Path
            path = Path(file_path)
            if not path.exists():
                self.logger.error(f"File not found: {file_path}")
                return False
            
            self.workbook = openpyxl.load_workbook(file_path)
            self.current_file_path = file_path
            self.logger.info(f"Workbook loaded for report generation: {file_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error loading workbook: {e}")
            return False
    
    def create_report(self, duplicates: List[DuplicateRecord]) -> bool:
        """
        Create a Duplicate Report worksheet.
        
        Args:
            duplicates: List of DuplicateRecord objects.
            
        Returns:
            True if successful, False otherwise.
        """
        if not self.workbook:
            self.logger.error("No workbook loaded")
            return False
        
        try:
            # Remove existing report sheet if it exists
            if self.REPORT_SHEET_NAME in self.workbook.sheetnames:
                del self.workbook[self.REPORT_SHEET_NAME]
            
            # Create new worksheet
            worksheet = self.workbook.create_sheet(self.REPORT_SHEET_NAME)
            self.logger.info(f"Created worksheet: {self.REPORT_SHEET_NAME}")
            
            # Set up headers
            headers = [
                "Batch Number",
                "Worksheet",
                "Row Number",
                "Cell Reference",
                "Occurrences",
                "Duplicate Type",
                "Description"
            ]
            
            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = self.THIN_BORDER
            
            # Write duplicate records
            for row_idx, duplicate in enumerate(duplicates, 2):
                worksheet.cell(row=row_idx, column=1).value = duplicate.batch_number
                worksheet.cell(row=row_idx, column=2).value = duplicate.worksheet
                worksheet.cell(row=row_idx, column=3).value = duplicate.row_number
                worksheet.cell(row=row_idx, column=4).value = duplicate.cell_reference
                worksheet.cell(row=row_idx, column=5).value = duplicate.occurrences
                worksheet.cell(row=row_idx, column=6).value = duplicate.duplicate_type
                
                # Generate description
                description = self._generate_description(duplicate)
                worksheet.cell(row=row_idx, column=7).value = description
                
                # Apply borders and alignment to data rows
                for col_idx in range(1, len(headers) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.border = self.THIN_BORDER
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Adjust column widths
            column_widths = [20, 25, 12, 18, 12, 18, 40]
            for col_idx, width in enumerate(column_widths, 1):
                worksheet.column_dimensions[get_column_letter(col_idx)].width = width
            
            # Freeze header row
            worksheet.freeze_panes = "A2"
            
            self.logger.info(f"Report created with {len(duplicates)} duplicate records")
            return True
            
        except Exception as e:
            self.logger.error(f"Error creating duplicate report: {e}")
            return False
    
    def _generate_description(self, duplicate: DuplicateRecord) -> str:
        """
        Generate a detailed description for a duplicate record.
        
        Args:
            duplicate: DuplicateRecord object.
            
        Returns:
            Formatted description string.
        """
        if duplicate.duplicate_type == "Same Cell":
            return (f"Batch '{duplicate.batch_number}' appears {duplicate.occurrences} times "
                   f"in the same cell at row {duplicate.row_number}")
        else:  # Different Rows
            return (f"Batch '{duplicate.batch_number}' appears in {duplicate.occurrences} "
                   f"different rows (one at row {duplicate.row_number})")
    
    def save_workbook(self, output_path: str = None) -> bool:
        """
        Save the workbook with the duplicate report.
        
        Args:
            output_path: Path to save the workbook. If None, overwrites original.
            
        Returns:
            True if successful, False otherwise.
        """
        if not self.workbook:
            self.logger.error("No workbook loaded")
            return False
        
        try:
            save_path = output_path or self.current_file_path
            self.workbook.save(save_path)
            self.logger.info(f"Workbook saved with duplicate report: {save_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error saving workbook: {e}")
            return False
    
    def close(self):
        """Close the workbook without saving."""
        if self.workbook:
            self.workbook.close()
            self.logger.debug("Workbook closed")
