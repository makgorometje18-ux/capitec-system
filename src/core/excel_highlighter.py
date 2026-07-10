"""
Excel Highlighter Module - Applies formatting to Excel cells based on validation results.
"""

from typing import List, Dict, Optional, Tuple
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.utils.logger import get_logger


class ExcelHighlighter:
    """
    Applies visual highlighting to Excel workbooks to indicate validation errors.
    
    Highlights errors in different colors:
    - Red: Critical errors (duplicates, blank fields, invalid types)
    - Yellow: Warnings (format issues, mismatches)
    - Green: Passed validation
    """
    
    # Color definitions
    ERROR_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
    ERROR_FONT = Font(color="FFFFFF", bold=True)  # White bold text
    
    WARNING_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
    WARNING_FONT = Font(color="000000", bold=True)  # Black bold text
    
    PASS_FILL = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")  # Green
    PASS_FONT = Font(color="FFFFFF", bold=True)  # White bold text
    
    def __init__(self):
        """Initialize the Excel Highlighter."""
        self.logger = get_logger()
        self.workbook = None
        self.current_file_path = None
    
    def load_workbook(self, file_path: str) -> bool:
        """
        Load a workbook for highlighting.
        
        Args:
            file_path: Path to the Excel workbook.
            
        Returns:
            True if successful, False otherwise.
        """
        try:
            path = Path(file_path)
            if not path.exists():
                self.logger.error(f"File not found: {file_path}")
                return False
            
            self.workbook = openpyxl.load_workbook(file_path)
            self.current_file_path = file_path
            self.logger.info(f"Workbook loaded for highlighting: {file_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error loading workbook for highlighting: {e}")
            return False
    
    def highlight_errors(self, sheet_name: str, errors: Dict[int, List[str]]) -> bool:
        """
        Apply error highlighting to rows based on validation errors.
        
        Args:
            sheet_name: Name of the worksheet to highlight.
            errors: Dictionary mapping row numbers to list of error messages.
            
        Returns:
            True if successful, False otherwise.
        """
        if not self.workbook:
            self.logger.error("No workbook loaded")
            return False
        
        try:
            worksheet = self.workbook[sheet_name]
            
            for row_num, error_list in errors.items():
                if row_num <= 1:  # Skip header row
                    continue
                
                # Highlight entire row in red for errors
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_num, column=col_idx)
                    cell.fill = self.ERROR_FILL
                    cell.font = self.ERROR_FONT
            
            self.logger.info(f"Applied error highlighting to {len(errors)} rows in {sheet_name}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error highlighting errors: {e}")
            return False
    
    def highlight_warnings(self, sheet_name: str, warnings: Dict[int, List[str]]) -> bool:
        """
        Apply warning highlighting to rows.
        
        Args:
            sheet_name: Name of the worksheet to highlight.
            warnings: Dictionary mapping row numbers to list of warning messages.
            
        Returns:
            True if successful, False otherwise.
        """
        if not self.workbook:
            self.logger.error("No workbook loaded")
            return False
        
        try:
            worksheet = self.workbook[sheet_name]
            
            for row_num, warning_list in warnings.items():
                if row_num <= 1:  # Skip header row
                    continue
                
                # Highlight entire row in yellow for warnings
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_num, column=col_idx)
                    cell.fill = self.WARNING_FILL
                    cell.font = self.WARNING_FONT
            
            self.logger.info(f"Applied warning highlighting to {len(warnings)} rows in {sheet_name}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error highlighting warnings: {e}")
            return False
    
    def highlight_cell(self, sheet_name: str, row: int, column: str, 
                      error_type: str = "error") -> bool:
        """
        Highlight a specific cell.
        
        Args:
            sheet_name: Name of the worksheet.
            row: Row number (1-indexed).
            column: Column letter (e.g., 'A', 'B').
            error_type: Type of highlighting ('error', 'warning', 'pass').
            
        Returns:
            True if successful, False otherwise.
        """
        if not self.workbook:
            self.logger.error("No workbook loaded")
            return False
        
        try:
            worksheet = self.workbook[sheet_name]
            cell = worksheet[f"{column}{row}"]
            
            if error_type == "error":
                cell.fill = self.ERROR_FILL
                cell.font = self.ERROR_FONT
            elif error_type == "warning":
                cell.fill = self.WARNING_FILL
                cell.font = self.WARNING_FONT
            elif error_type == "pass":
                cell.fill = self.PASS_FILL
                cell.font = self.PASS_FONT
            
            return True
            
        except Exception as e:
            self.logger.error(f"Error highlighting cell {column}{row}: {e}")
            return False
    
    def highlight_duplicate_rows(self, sheet_name: str, duplicate_rows: List[int]) -> bool:
        """
        Highlight rows containing duplicates.
        
        Args:
            sheet_name: Name of the worksheet.
            duplicate_rows: List of row numbers with duplicates.
            
        Returns:
            True if successful, False otherwise.
        """
        if not self.workbook:
            self.logger.error("No workbook loaded")
            return False
        
        try:
            worksheet = self.workbook[sheet_name]
            
            for row_num in duplicate_rows:
                if row_num <= 1:  # Skip header
                    continue
                
                # Highlight Batch_No column specifically for duplicates
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_num, column=col_idx)
                    cell.fill = self.ERROR_FILL
                    cell.font = self.ERROR_FONT
            
            self.logger.info(f"Highlighted {len(duplicate_rows)} duplicate rows in {sheet_name}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error highlighting duplicates: {e}")
            return False
    
    def save_workbook(self, output_path: Optional[str] = None) -> bool:
        """
        Save the workbook with highlighting applied.
        
        Args:
            output_path: Path to save the workbook. If None, overwrites original.
            
        Returns:
            True if successful, False otherwise.
        """
        if not self.workbook:
            self.logger.error("No workbook loaded")
            return False
        
        try:
            save_path = output_path or self.current_file_path
            self.workbook.save(save_path)
            self.logger.info(f"Workbook saved with highlighting: {save_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error saving workbook: {e}")
            return False
    
    def close(self):
        """Close the workbook without saving."""
        if self.workbook:
            self.workbook.close()
            self.logger.debug("Workbook closed")
