"""
Sample Workbook Generator - Creates test Excel files for validation testing.

This utility generates sample Excel workbooks with various test scenarios
to demonstrate the validation engine.
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime


def create_valid_daily_output_workbook() -> Workbook:
    """
    Create a sample valid Daily Output File workbook.
    
    Returns:
        Openpyxl Workbook with valid test data.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "DAILY OUTPUT FILE 03-07-2026"
    
    # Create headers
    headers = [
        'Order_no', 'Order_Creation_Date', 'Branch_Code', 'Branch_Name', 'Card_Type',
        'Number_of_Batches', 'Waybill_Number', 'Batch_Number', 'BagNumber'
    ]
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Add valid test data
    test_data = [
        [1, datetime.now().date(), 'BR001', 'Branch Johannesburg', 'SIM', 2, 'WB001', '10001|10002', "'00001234"],
        [2, datetime.now().date(), 'BR002', 'Branch Pretoria', 'DMCCLS', 3, 'WB002', '20001|20002|20003', "'00002345"],
        [3, datetime.now().date(), 'BR003', 'Branch Cape Town', 'SIM', 1, 'WB003', '30001', "'00003456"],
    ]
    
    for row_data in test_data:
        ws.append(row_data)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb


def create_invalid_daily_output_workbook() -> Workbook:
    """
    Create a sample invalid Daily Output File workbook with various errors.
    
    Returns:
        Openpyxl Workbook with test data containing errors.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "DAILY OUTPUT FILE 03-07-2026"
    
    # Create headers
    headers = [
        'Order_no', 'Order_Creation_Date', 'Branch_Code', 'Branch_Name', 'Card_Type',
        'Number_of_Batches', 'Waybill_Number', 'Batch_Number', 'BagNumber'
    ]
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Add invalid test data with various errors
    test_data = [
        # Row 2: Blank Branch_Code
        [1, datetime.now().date(), '', 'Branch Johannesburg', 'SIM', 2, 'WB001', '10001|10002', "'00001234"],
        
        # Row 3: Duplicate batch number in same cell (10001 appears twice)
        [2, datetime.now().date(), 'BR002', 'Branch Pretoria', 'DMCCLS', 2, 'WB002', '20001|20001', "'00002345"],
        
        # Row 4: Number_of_Batches mismatch (says 3 but only has 2 batches)
        [3, datetime.now().date(), 'BR003', 'Branch Cape Town', 'SIM', 3, 'WB003', '30001|30002', "'00003456"],
        
        # Row 5: Invalid BagNumber format (double apostrophe)
        [4, datetime.now().date(), 'BR004', 'Branch Durban', 'SIM', 1, 'WB004', '40001', "''00004567"],
        
        # Row 6: Blank Batch_Number
        [5, datetime.now().date(), 'BR005', 'Branch Bloemfontein', 'DMCCLS', 1, 'WB005', '', "'00005678"],
        
        # Row 7: Duplicate across rows (30001 also in row 4)
        [6, datetime.now().date(), 'BR006', 'Branch Port Elizabeth', 'SIM', 1, 'WB006', '30001', "'00006789"],
    ]
    
    for row_data in test_data:
        ws.append(row_data)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb


def generate_sample_workbooks():
    """
    Generate sample workbooks for testing.
    
    Creates:
    1. sample_valid.xlsx - Valid data for successful validation
    2. sample_invalid.xlsx - Invalid data with various errors
    """
    sample_dir = Path("sample_files")
    sample_dir.mkdir(exist_ok=True)
    
    print("Generating sample workbooks...\n")
    
    # Generate valid workbook
    valid_wb = create_valid_daily_output_workbook()
    valid_path = sample_dir / "sample_valid.xlsx"
    valid_wb.save(valid_path)
    print(f"✓ Created: {valid_path}")
    
    # Generate invalid workbook
    invalid_wb = create_invalid_daily_output_workbook()
    invalid_path = sample_dir / "sample_invalid.xlsx"
    invalid_wb.save(invalid_path)
    print(f"✓ Created: {invalid_path}")
    
    print("\nSample workbooks created successfully!")
    print(f"Use these files to test the Validation Engine:\n")
    print(f"  Valid:   {valid_path}")
    print(f"  Invalid: {invalid_path}")


if __name__ == "__main__":
    generate_sample_workbooks()
