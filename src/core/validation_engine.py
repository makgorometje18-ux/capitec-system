"""
Comprehensive Validation Engine - Orchestrates all validation checks.
"""

from typing import List, Optional, Dict, Any, Tuple, Callable
from datetime import datetime
from pathlib import Path
import time

from src.models.models import (
    Workbook, ValidationResult, ValidationError, DuplicateRecord, ErrorSummary
)

# Mapping of internal error types to display names
ERROR_TYPE_DISPLAY = {
    "DUPLICATE_BATCH": "Duplicate Batch Number",
    "DUPLICATE_BATCH_SAME_CELL": "Duplicate Within Cell",
    "DUPLICATE_ACROSS_ROWS": "Duplicate Across Rows",
    "DUPLICATE_CROSS_WORKBOOK": "Cross Workbook Duplicate",
    "BATCH_MISMATCH": "Invalid Number of Batches",
    "INVALID_BAG": "Invalid Bag Number",
    "BLANK_FIELD": "Blank Field",
    "MISSING_HEADER": "Missing Header",
    "INVALID_CARD_TYPE": "Invalid Card Type",
    "SYSTEM_ERROR": "Unknown"
}
from src.core.workbook_loader import WorkbookLoader
from src.core.excel_highlighter import ExcelHighlighter
from src.core.duplicate_report import DuplicateReportGenerator
from src.core.auto_fixer import AutoFixer
from src.core.error_summary_builder import ErrorSummaryBuilder
from src.core.cross_workbook_duplicate_checker import CrossWorkbookDuplicateChecker
from src.utils.logger import get_logger
from src.utils.helpers import split_batch_numbers, validate_bag_number, clean_string
from openpyxl.utils import get_column_letter

try:
    from src.core.audit_manager import AuditManager
    HAS_AUDIT = True
except ImportError:
    HAS_AUDIT = False


class ValidationEngine:
    """
    Comprehensive validation engine that orchestrates all validation checks.
    
    Coordinates workbook loading, sheet detection, header validation,
    duplicate checking, batch validation, bag validation, and blank
    field validation according to business rules.
    """

    REQUIRED_HEADERS = WorkbookLoader.REQUIRED_HEADERS
    MANDATORY_FIELDS = REQUIRED_HEADERS

    def __init__(self, progress_callback: Optional[Callable[[int, str, str], None]] = None) -> None:
        """
        Initialize the Validation Engine.
        
        Args:
            progress_callback: Optional function called after each validation stage.
                Signature: callback(percent: int, stage: str, status: str)
        """
        self.logger = get_logger()
        self.workbook_loader = WorkbookLoader()
        self.duplicates: List[DuplicateRecord] = []
        self.validation_errors: List[ValidationError] = []
        self.audit_manager = AuditManager() if HAS_AUDIT else None
        self.progress_callback = progress_callback

    def _emit_progress(self, percent: int, stage: str, status: str = "running") -> None:
        """Emit progress update if a callback is registered."""
        if self.progress_callback:
            try:
                self.progress_callback(percent, stage, status)
            except Exception:
                pass

    def _get_cell_reference(self, sheet_name: str, row_num: int, column_name: str) -> str:
        """
        Get the cell reference (e.g., 'B2', 'C5') for a given row and column name.
        
        Args:
            sheet_name: Name of the worksheet.
            row_num: Row number (1-based).
            column_name: Column header name.
            
        Returns:
            Cell reference string like 'B2', or '?' if not found.
        """
        try:
            headers = self.workbook_loader.get_headers(sheet_name)
            if not headers:
                return f"{column_name}_{row_num}"
            
            # Find column index (1-based)
            col_idx = None
            for idx, header in enumerate(headers, start=1):
                if clean_string(header) == clean_string(column_name):
                    col_idx = idx
                    break
            
            if col_idx is None:
                return f"{column_name}_{row_num}"
            
            # Convert to column letter
            col_letter = get_column_letter(col_idx)
            return f"{col_letter}{row_num}"
            
        except Exception as e:
            self.logger.debug(f"Could not get cell reference for {column_name} at row {row_num}: {e}")
            return f"{column_name}_{row_num}"

    def validate_complete_workbook(self, file_path: str, previous_folder: Optional[str] = None) -> ValidationResult:
        """
        Validate a complete workbook end-to-end.
        
        Args:
            file_path: Path to the Excel workbook.
            
        Returns:
            ValidationResult with all validation details.
        """
        result = ValidationResult()
        start_time = time.time()
        
        try:
            self.logger.info(f"Starting validation for: {file_path}")
            
            # Step 1: Load workbook
            self._emit_progress(3, "Loading workbook", "running")
            workbook = self.workbook_loader.load_workbook(file_path)
            if not workbook:
                error_msg = "Failed to load workbook"
                result.errors.append(error_msg)
                result.validation_errors.append(ValidationError(
                    rule_id="WORKBOOK_LOAD_FAILED",
                    error_type="SYSTEM_ERROR",
                    worksheet="",
                    row_number=0,
                    column_name="",
                    cell_reference="",
                    error_message=error_msg
                ))
                result.error_count = 1
                result.passed = False
                self._emit_progress(100, "Failed to load workbook", "failed")
                return result
            
            # Step 2: Detect Daily Output sheet
            self._emit_progress(10, "Detecting Daily Output worksheet", "running")
            daily_output_sheet = self.workbook_loader.detect_daily_output_sheet()
            if not daily_output_sheet:
                error_msg = "Daily Output File worksheet not found"
                result.errors.append(error_msg)
                result.validation_errors.append(ValidationError(
                    rule_id="DAILY_OUTPUT_NOT_FOUND",
                    error_type="SYSTEM_ERROR",
                    worksheet="",
                    row_number=0,
                    column_name="",
                    cell_reference="",
                    error_message=error_msg
                ))
                result.error_count += 1
                self._emit_progress(100, "Daily Output sheet not found", "failed")
            else:
                self.logger.info(f"Using daily worksheet for validation: {daily_output_sheet}")

                worksheet = self.workbook_loader.get_worksheet(daily_output_sheet)
                if worksheet:
                    all_rows = list(worksheet.iter_rows(min_row=1, values_only=True))
                    total_raw_rows = len(all_rows)
                    non_empty_rows = sum(
                        1 for row in all_rows
                        if any(cell is not None and str(cell).strip() != '' for cell in row)
                    )
                    self.logger.info(
                        "Selected sheet '%s' max_row=%s total_raw_rows=%d non_empty_rows=%d",
                        daily_output_sheet,
                        worksheet.max_row,
                        total_raw_rows,
                        non_empty_rows,
                    )
                    self.logger.info("Sample first 5 raw rows: %s", all_rows[:5])

                # Count data rows (rows after header row, with content)
                data_rows = self.workbook_loader.get_data_rows(daily_output_sheet)
                if data_rows is not None:
                    result.rows_processed = len(data_rows)
                    self.logger.info(f"Rows processed (data rows): {result.rows_processed}")
                else:
                    result.rows_processed = 0
                    self.logger.warning("Could not get data rows, setting rows_processed=0")

                # Step 3: Validate headers
                self._emit_progress(20, "Validating worksheet structure", "running")
                header_result = self._validate_headers(daily_output_sheet)
                result.errors.extend(header_result['errors'])
                result.error_count += len(header_result['errors'])
                result.validation_errors.extend(header_result.get('validation_errors', []))
                
                if header_result['errors']:
                    self._emit_progress(100, "Header validation failed", "failed")
                    result.duration_seconds = time.time() - start_time
                    result.passed = False
                    return result
                
                # Step 4: Validate duplicates
                self._emit_progress(30, "Checking duplicate batches", "running")
                dup_result = self._validate_duplicates(daily_output_sheet)
                self.duplicates = dup_result['duplicates']
                result.errors.extend(dup_result['errors'])
                result.error_count += len(dup_result['errors'])
                result.validation_errors.extend(dup_result.get('validation_errors', []))

                # Step 4b: Cross-workbook duplicates (optional)
                if previous_folder:
                    self._emit_progress(42, "Checking cross-workbook duplicates", "running")
                    cross_result = self._validate_cross_workbook_duplicates(daily_output_sheet, previous_folder)
                    self.duplicates.extend(cross_result.get('duplicates', []))
                    result.errors.extend(cross_result.get('errors', []))
                    result.error_count += len(cross_result.get('errors', []))
                    result.validation_errors.extend(cross_result.get('validation_errors', []))
                
                # Step 5: Validate batch counts
                self._emit_progress(52, "Validating Number_of_Batches", "running")
                batch_result = self._validate_batch_counts(daily_output_sheet)
                result.errors.extend(batch_result['errors'])
                result.error_count += len(batch_result['errors'])
                result.validation_errors.extend(batch_result.get('validation_errors', []))
                
                # Step 6: Validate bag numbers
                self._emit_progress(62, "Validating Bag Numbers", "running")
                bag_result = self._validate_bag_numbers(daily_output_sheet)
                result.errors.extend(bag_result['errors'])
                result.error_count += len(bag_result['errors'])
                result.validation_errors.extend(bag_result.get('validation_errors', []))
                
                # Step 7: Validate blank fields
                self._emit_progress(72, "Checking blank fields", "running")
                blank_result = self._validate_blank_fields(daily_output_sheet)
                result.errors.extend(blank_result['errors'])
                result.error_count += len(blank_result['errors'])
                result.validation_errors.extend(blank_result.get('validation_errors', []))
                
                # Step 8: Validate card types
                self._emit_progress(82, "Validating Card Types", "running")
                card_result = self._validate_card_types(daily_output_sheet)
                result.errors.extend(card_result['errors'])
                result.error_count += len(card_result['errors'])
                result.validation_errors.extend(card_result.get('validation_errors', []))
            
            # Determine pass/fail
            result.passed = result.error_count == 0
            result.duration_seconds = time.time() - start_time
            
            if result.passed:
                self._emit_progress(100, "Validation completed successfully", "complete")
            else:
                self._emit_progress(100, "Validation completed with errors", "complete")
            
            self.logger.info(f"Validation completed. "
                           f"Passed: {result.passed}, "
                           f"Errors: {result.error_count}, "
                           f"Duration: {result.duration_seconds:.2f}s")
            
            return result
            
        except Exception as e:
            self.logger.error(f"Validation engine error: {e}")
            error_msg = f"Validation engine error: {str(e)}"
            result.errors.append(error_msg)
            result.validation_errors.append(ValidationError(
                rule_id="VALIDATION_ENGINE_ERROR",
                error_type="SYSTEM_ERROR",
                worksheet="",
                row_number=0,
                column_name="",
                cell_reference="",
                error_message=error_msg
            ))
            result.error_count += 1
            result.passed = False
            result.duration_seconds = time.time() - start_time
            self._emit_progress(100, f"Validation failed: {str(e)[:50]}", "failed")
            return result
        
        finally:
            self.workbook_loader.close()

    def _validate_headers(self, sheet_name: str) -> Dict[str, Any]:
        """
        Validate that required headers are present.
        
        Business Rule: All required headers must be present.
        
        Args:
            sheet_name: Name of the worksheet to validate.
            
        Returns:
            Dictionary with errors list and structured validation errors.
        """
        self.logger.info(f"Validating headers in {sheet_name}")
        errors = []
        validation_errors = []
        
        headers = self.workbook_loader.get_headers(sheet_name)
        if not headers:
            error_msg = "Failed to read headers from worksheet"
            errors.append(error_msg)
            validation_errors.append(ValidationError(
                rule_id="HEADER_READ_FAILED",
                error_type="SYSTEM_ERROR",
                worksheet=sheet_name,
                row_number=0,
                column_name="",
                cell_reference="",
                error_message=error_msg
            ))
            return {'errors': errors, 'validation_errors': validation_errors}
        
        # Check for missing required headers
        header_set = set(headers)
        missing_headers = self.MANDATORY_FIELDS - header_set
        
        if missing_headers:
            error_msg = f"Missing required headers: {', '.join(missing_headers)}"
            errors.append(error_msg)
            validation_errors.append(ValidationError(
                rule_id="MISSING_HEADERS",
                error_type="MISSING_HEADER",
                worksheet=sheet_name,
                row_number=0,
                column_name=", ".join(missing_headers),
                cell_reference="",
                error_message=error_msg
            ))
            self.logger.warning(error_msg)
        else:
            self.logger.info(f"All required headers present: {self.MANDATORY_FIELDS}")
        
        return {'errors': errors, 'validation_errors': validation_errors}

    def _validate_duplicates(self, sheet_name: str) -> Dict[str, Any]:
        """
        Validate for duplicate batch numbers.
        
        Business Rule BR001: Batch numbers must be unique across the worksheet.
        Business Rule BR002: Batch numbers must not repeat inside a single cell.
        Business Rule BR003: Batch_Number values are separated using '|'.
        
        Args:
            sheet_name: Name of the worksheet to validate.
            
        Returns:
            Dictionary with duplicates, errors, and structured validation errors.
        """
        self.logger.info(f"Validating duplicates in {sheet_name}")
        errors = []
        validation_errors = []
        duplicates = []
        all_batches = {}  # batch_number -> [(row, cell_ref, duplicate_type)]
        
        try:
            batch_values = self.workbook_loader.get_column_values(sheet_name, 'Batch_Number')
            if not batch_values:
                return {'errors': errors, 'duplicates': duplicates, 'validation_errors': validation_errors}
            
            # Check each batch_no cell
            for row_num, batch_string in batch_values:
                if not batch_string:
                    continue
                
                # Split batches from this cell
                batches = split_batch_numbers(batch_string)
                
                # Check for duplicates within same cell (BR002)
                if len(batches) != len(set(batches)):
                    duplicate_batches = set()
                    for batch in batches:
                        if batches.count(batch) > 1:
                            duplicate_batches.add(batch)
                    
                    for batch in duplicate_batches:
                        cell_ref = self._get_cell_reference(sheet_name, row_num, 'Batch_Number')
                        error_msg = (f"Duplicate batch '{batch}' found in same cell at "
                                   f"row {row_num}, cell {cell_ref}")
                        errors.append(error_msg)
                        validation_errors.append(ValidationError(
                            rule_id="DUPLICATE_BATCH_SAME_CELL",
                            error_type="DUPLICATE_BATCH_SAME_CELL",
                            worksheet=sheet_name,
                            row_number=row_num,
                            column_name="Batch_Number",
                            cell_reference=cell_ref,
                            error_message=error_msg,
                            invalid_value=batch
                        ))
                        self.logger.warning(error_msg)
                
                # Track all batches across rows
                for batch in batches:
                    if batch not in all_batches:
                        all_batches[batch] = []
                    cell_ref = self._get_cell_reference(sheet_name, row_num, 'Batch_Number')
                    all_batches[batch].append((row_num, f"Batch_Number", cell_ref))
            
            # Check for duplicates across rows (BR001)
            for batch_number, occurrences in all_batches.items():
                if len(occurrences) > 1:
                    cell_refs = [cell_ref for _, _, cell_ref in occurrences]
                    unique_cells = sorted(set(cell_refs))
                    cell_ref_str = ", ".join(unique_cells)
                    
                    # Determine duplicate type based on whether all occurrences are in the same cell
                    duplicate_type = "Same Cell" if len(unique_cells) == 1 else "Different Rows"
                    error_msg = f"Duplicate batch number '{batch_number}' found in {len(occurrences)} rows at cells: {cell_ref_str}"
                    errors.append(error_msg)
                    self.logger.warning(error_msg)
                    
                    # Create duplicate records for each occurrence
                    for row_num, col, cell_ref in occurrences:
                        duplicate = DuplicateRecord(
                            batch_number=batch_number,
                            worksheet=sheet_name,
                            row_number=row_num,
                            cell_reference=f"Batch_Number_{row_num}",
                            occurrences=len(occurrences),
                            duplicate_type=duplicate_type
                        )
                        duplicates.append(duplicate)
                    
                    # Create ONE validation error for the first occurrence only
                    first_row, _, _ = occurrences[0]
                    first_cell_ref = self._get_cell_reference(sheet_name, first_row, 'Batch_Number')
                    validation_errors.append(ValidationError(
                        rule_id="DUPLICATE_BATCH_NUMBER",
                        error_type="DUPLICATE_ACROSS_ROWS" if duplicate_type == "Different Rows" else "DUPLICATE_BATCH_SAME_CELL",
                        worksheet=sheet_name,
                        row_number=first_row,
                        column_name="Batch_Number",
                        cell_reference=first_cell_ref,
                        error_message=error_msg,
                        invalid_value=batch_number
                    ))
            
            if not errors:
                self.logger.info("No duplicate batch numbers found")
            
            return {'errors': errors, 'duplicates': duplicates, 'validation_errors': validation_errors}
            
        except Exception as e:
            error_msg = f"Error validating duplicates: {e}"
            errors.append(error_msg)
            validation_errors.append(ValidationError(
                rule_id="DUPLICATE_VALIDATION_ERROR",
                error_type="SYSTEM_ERROR",
                worksheet=sheet_name,
                row_number=0,
                column_name="",
                cell_reference="",
                error_message=error_msg
            ))
            self.logger.error(error_msg)
            return {'errors': errors, 'duplicates': duplicates, 'validation_errors': validation_errors}

    def _validate_batch_counts(self, sheet_name: str) -> Dict[str, Any]:
        """
        Validate that Number_of_Batches matches the number of batches in Batch_Number.
        
        Business Rule BR004: Number_of_Batches must equal the number of values found in Batch_Number.
        
        Args:
            sheet_name: Name of the worksheet to validate.
            
        Returns:
            Dictionary with errors list and structured validation errors.
        """
        self.logger.info(f"Validating batch counts in {sheet_name}")
        errors = []
        validation_errors = []
        
        try:
            data_rows = self.workbook_loader.get_data_rows(sheet_name)
            if not data_rows:
                return {'errors': errors, 'validation_errors': validation_errors}
            
            for row_data in data_rows:
                row_num = row_data.get('Batch_Number', {}).get('row', 0)
                
                # Get Number_of_Batches value
                no_of_batches_data = row_data.get('Number_of_Batches', {})
                no_of_batches_value = no_of_batches_data.get('value')
                
                # Get Batch_Number value
                batch_no_data = row_data.get('Batch_Number', {})
                batch_no_value = batch_no_data.get('value')
                
                if not batch_no_value or not no_of_batches_value:
                    continue
                
                try:
                    expected_count = int(no_of_batches_value)
                    batches = split_batch_numbers(str(batch_no_value))
                    actual_count = len(batches)
                    
                    if expected_count != actual_count:
                        cell_ref = self._get_cell_reference(sheet_name, row_num, 'Number_of_Batches')
                        error_msg = (f"Row {row_num}, cell {cell_ref}: Number_of_Batches mismatch. "
                                   f"Expected {expected_count}, found {actual_count}")
                        errors.append(error_msg)
                        validation_errors.append(ValidationError(
                            rule_id="BATCH_COUNT_MISMATCH",
                            error_type="BATCH_MISMATCH",
                            worksheet=sheet_name,
                            row_number=row_num,
                            column_name="Number_of_Batches",
                            cell_reference=cell_ref,
                            error_message=error_msg,
                            invalid_value=f"Expected {expected_count}, found {actual_count}"
                        ))
                        self.logger.warning(error_msg)
                
                except ValueError:
                    cell_ref = self._get_cell_reference(sheet_name, row_num, 'Number_of_Batches')
                    error_msg = f"Row {row_num}, cell {cell_ref}: Invalid Number_of_Batches value: {no_of_batches_value}"
                    errors.append(error_msg)
                    validation_errors.append(ValidationError(
                        rule_id="INVALID_BATCH_COUNT",
                        error_type="BATCH_MISMATCH",
                        worksheet=sheet_name,
                        row_number=row_num,
                        column_name="Number_of_Batches",
                        cell_reference=cell_ref,
                        error_message=error_msg,
                        invalid_value=str(no_of_batches_value)
                    ))
                    self.logger.warning(error_msg)
            
            if not errors:
                self.logger.info("All batch counts are valid")
            
            return {'errors': errors, 'validation_errors': validation_errors}
            
        except Exception as e:
            error_msg = f"Error validating batch counts: {e}"
            errors.append(error_msg)
            validation_errors.append(ValidationError(
                rule_id="BATCH_COUNT_VALIDATION_ERROR",
                error_type="SYSTEM_ERROR",
                worksheet=sheet_name,
                row_number=0,
                column_name="",
                cell_reference="",
                error_message=error_msg
            ))
            self.logger.error(error_msg)
            return {'errors': errors, 'validation_errors': validation_errors}

    def _validate_bag_numbers(self, sheet_name: str) -> Dict[str, Any]:
        """
        Validate bag number formatting.
        
        Business Rule BR005: BagNumber must contain exactly one leading apostrophe
        or be a plain numeric value.
        
        Args:
            sheet_name: Name of the worksheet to validate.
            
        Returns:
            Dictionary with errors list and structured validation errors.
        """
        self.logger.info(f"Validating bag numbers in {sheet_name}")
        errors = []
        validation_errors = []
        
        try:
            data_rows = self.workbook_loader.get_data_rows(sheet_name)
            if data_rows is None:
                return {'errors': errors, 'validation_errors': validation_errors}
            
            for row_data in data_rows:
                bag_data = row_data.get('BagNumber', {})
                row_num = bag_data.get('row') or row_data.get('Batch_Number', {}).get('row', 0)
                bag_no_value = bag_data.get('value')

                if bag_no_value is None or str(bag_no_value).strip() == '':
                    cell_ref = self._get_cell_reference(sheet_name, row_num, 'BagNumber')
                    error_msg = f"Row {row_num}, cell {cell_ref}: Blank BagNumber value"
                    errors.append(error_msg)
                    validation_errors.append(ValidationError(
                        rule_id="BLANK_BAG_NUMBER",
                        error_type="INVALID_BAG",
                        worksheet=sheet_name,
                        row_number=row_num,
                        column_name="BagNumber",
                        cell_reference=cell_ref,
                        error_message=error_msg,
                        invalid_value=""
                    ))
                    self.logger.warning(error_msg)
                    continue

                # Support multiple bag numbers separated by |
                bag_numbers = split_batch_numbers(str(bag_no_value))
                invalid_bags = []
                for bag in bag_numbers:
                    if not validate_bag_number(bag):
                        invalid_bags.append(bag)
                
                if invalid_bags:
                    cell_ref = self._get_cell_reference(sheet_name, row_num, 'BagNumber')
                    invalid_str = ", ".join(invalid_bags)
                    error_msg = (f"Row {row_num}, cell {cell_ref}: Invalid BagNumber format '{invalid_str}'. "
                               f"Must be numeric or start with exactly one apostrophe followed by digits.")
                    errors.append(error_msg)
                    validation_errors.append(ValidationError(
                        rule_id="INVALID_BAG_FORMAT",
                        error_type="INVALID_BAG",
                        worksheet=sheet_name,
                        row_number=row_num,
                        column_name="BagNumber",
                        cell_reference=cell_ref,
                        error_message=error_msg,
                        invalid_value=invalid_str
                    ))
                    self.logger.warning(error_msg)
            
            if not errors:
                self.logger.info("All bag numbers are valid")
            
            return {'errors': errors, 'validation_errors': validation_errors}
            
        except Exception as e:
            error_msg = f"Error validating bag numbers: {e}"
            errors.append(error_msg)
            validation_errors.append(ValidationError(
                rule_id="BAG_VALIDATION_ERROR",
                error_type="SYSTEM_ERROR",
                worksheet=sheet_name,
                row_number=0,
                column_name="",
                cell_reference="",
                error_message=error_msg
            ))
            self.logger.error(error_msg)
            return {'errors': errors, 'validation_errors': validation_errors}

    def _validate_blank_fields(self, sheet_name: str) -> Dict[str, Any]:
        """
        Validate that mandatory fields are not blank.
        
        Args:
            sheet_name: Name of the worksheet to validate.
            
        Returns:
            Dictionary with errors list and structured validation errors.
        """
        self.logger.info(f"Validating blank fields in {sheet_name}")
        errors = []
        validation_errors = []
        
        try:
            data_rows = self.workbook_loader.get_data_rows(sheet_name)
            if not data_rows:
                return {'errors': errors, 'validation_errors': validation_errors}
            
            for row_data in data_rows:
                for field_name in self.MANDATORY_FIELDS:
                    field_data = row_data.get(field_name, {})
                    row_num = field_data.get('row') or row_data.get('Batch_Number', {}).get('row', 0)
                    value = field_data.get('value')
                    
                    if value is None or clean_string(str(value)) == '':
                        cell_ref = self._get_cell_reference(sheet_name, row_num, field_name)
                        error_msg = f"Row {row_num}, cell {cell_ref}: Blank field '{field_name}'"
                        errors.append(error_msg)
                        validation_errors.append(ValidationError(
                            rule_id="BLANK_MANDATORY_FIELD",
                            error_type="BLANK_FIELD",
                            worksheet=sheet_name,
                            row_number=row_num,
                            column_name=field_name,
                            cell_reference=cell_ref,
                            error_message=error_msg,
                            invalid_value=""
                        ))
                        self.logger.warning(error_msg)
            
            if not errors:
                self.logger.info("No blank mandatory fields found")
            
            return {'errors': errors, 'validation_errors': validation_errors}
            
        except Exception as e:
            error_msg = f"Error validating blank fields: {e}"
            errors.append(error_msg)
            validation_errors.append(ValidationError(
                rule_id="BLANK_FIELD_VALIDATION_ERROR",
                error_type="SYSTEM_ERROR",
                worksheet=sheet_name,
                row_number=0,
                column_name="",
                cell_reference="",
                error_message=error_msg
            ))
            self.logger.error(error_msg)
            return {'errors': errors, 'validation_errors': validation_errors}

    def _validate_card_types(self, sheet_name: str) -> Dict[str, Any]:
        """
        Validate card types - only SIM or DMCCLS allowed.
        
        Business Rule BR009: Only SIM and DMCCLS are valid card types.
        
        Args:
            sheet_name: Name of the worksheet to validate.
            
        Returns:
            Dictionary with errors list and structured validation errors.
        """
        self.logger.info(f"Validating card types in {sheet_name}")
        errors = []
        validation_errors = []
        valid_card_types = {'SIM', 'DMCCLS'}
        
        try:
            card_values = self.workbook_loader.get_column_values(sheet_name, 'Card_Type')
            if not card_values:
                return {'errors': errors, 'validation_errors': validation_errors}
            
            for row_num, card_type_value in card_values:
                if not card_type_value:
                    continue
                
                card_type_cleaned = clean_string(str(card_type_value))
                if card_type_cleaned not in valid_card_types:
                    cell_ref = self._get_cell_reference(sheet_name, row_num, 'Card_Type')
                    error_msg = (f"Row {row_num}, cell {cell_ref}: Invalid Card_Type '{card_type_value}'. "
                               f"Must be either 'SIM' or 'DMCCLS'")
                    errors.append(error_msg)
                    validation_errors.append(ValidationError(
                        rule_id="INVALID_CARD_TYPE",
                        error_type="INVALID_CARD_TYPE",
                        worksheet=sheet_name,
                        row_number=row_num,
                        column_name="Card_Type",
                        cell_reference=cell_ref,
                        error_message=error_msg,
                        invalid_value=str(card_type_value)
                    ))
                    self.logger.warning(error_msg)
            
            if not errors:
                self.logger.info("All card types are valid")
            
            return {'errors': errors, 'validation_errors': validation_errors}
            
        except Exception as e:
            error_msg = f"Error validating card types: {e}"
            errors.append(error_msg)
            validation_errors.append(ValidationError(
                rule_id="CARD_TYPE_VALIDATION_ERROR",
                error_type="SYSTEM_ERROR",
                worksheet=sheet_name,
                row_number=0,
                column_name="",
                cell_reference="",
                error_message=error_msg
            ))
            self.logger.error(error_msg)
            return {'errors': errors, 'validation_errors': validation_errors}
    
    def generate_error_summary(self) -> ErrorSummary:
        """
        Generate an error summary from the last validation run.
        
        Returns:
            ErrorSummary object with categorized error counts.
        """
        builder = ErrorSummaryBuilder()
        return builder
    
    def highlight_errors_in_workbook(self, file_path: str, output_path: Optional[str] = None) -> bool:
        """
        Apply highlighting to errors in the workbook.
        
        Args:
            file_path: Path to the Excel workbook.
            output_path: Path to save the highlighted workbook. If None, overwrites original.
            
        Returns:
            True if successful, False otherwise.
        """
        try:
            highlighter = ExcelHighlighter()
            if not highlighter.load_workbook(file_path):
                return False
            
            # Get the daily output sheet name
            self.workbook_loader.load_workbook(file_path)
            sheet_name = self.workbook_loader.detect_daily_output_sheet()
            self.workbook_loader.close()
            
            if not sheet_name:
                self.logger.warning("Daily Output sheet not found for highlighting")
                return False
            
            # Organize errors by row
            errors_by_row: Dict[int, List[str]] = {}
            for error in self.validation_errors:
                row = error.row_number
                if row not in errors_by_row:
                    errors_by_row[row] = []
                errors_by_row[row].append(error.error_message)
            
            # Apply highlighting
            if errors_by_row:
                highlighter.highlight_errors(sheet_name, errors_by_row)
            
            # Save workbook
            if not highlighter.save_workbook(output_path):
                return False
            
            highlighter.close()
            self.logger.info("Workbook highlighting completed successfully")
            return True
            
        except Exception as e:
            self.logger.error(f"Error applying highlighting: {e}")
            return False
    
    def generate_duplicate_report(self, file_path: str, output_path: Optional[str] = None) -> bool:
        """
        Generate a Duplicate Report worksheet in the workbook.
        
        Creates a new worksheet "Duplicate Report" with detailed information
        about each duplicate batch number found.
        
        Args:
            file_path: Path to the Excel workbook.
            output_path: Path to save the workbook with report. If None, overwrites original.
            
        Returns:
            True if successful, False otherwise.
        """
        try:
            if not self.duplicates:
                self.logger.info("No duplicates to report")
                return True
            
            generator = DuplicateReportGenerator()
            if not generator.load_workbook(file_path):
                return False
            
            # Create the report
            if not generator.create_report(self.duplicates):
                return False
            
            # Save workbook
            if not generator.save_workbook(output_path):
                return False
            
            generator.close()
            self.logger.info(f"Duplicate report generated with {len(self.duplicates)} records")
            return True
            
        except Exception as e:
            self.logger.error(f"Error generating duplicate report: {e}")
            return False
    
    def auto_fix_workbook(self, file_path: str, output_path: Optional[str] = None) -> Tuple[bool, int]:
        """
        Auto-fix common issues in the workbook.
        
        Currently fixes:
        - Extra leading apostrophes in BagNumber values
        
        Args:
            file_path: Path to the Excel workbook.
            output_path: Path to save the fixed workbook. If None, overwrites original.
            
        Returns:
            Tuple of (success: bool, changes_made: int)
        """
        try:
            fixer = AutoFixer()
            if not fixer.load_workbook(file_path):
                return False, 0
            
            # Detect sheet name
            self.workbook_loader.load_workbook(file_path)
            sheet_name = self.workbook_loader.detect_daily_output_sheet()
            self.workbook_loader.close()
            
            if not sheet_name:
                self.logger.warning("Daily Output sheet not found for auto-fixing")
                return False, 0
            
            # Apply fixes
            success, changes = fixer.fix_bag_numbers(sheet_name)
            if not success:
                return False, 0
            
            # Save workbook
            if changes > 0:
                if not fixer.save_workbook(output_path):
                    return False, changes
            
            fixer.close()
            self.logger.info(f"Auto-fix completed: {changes} changes made")
            return True, changes
            
        except Exception as e:
            self.logger.error(f"Error in auto-fix: {e}")
            return False, 0

    def get_duplicates(self) -> List[DuplicateRecord]:
        """
        Get the list of duplicates found during validation.
        
        Returns:
            List of DuplicateRecord objects.
        """
        return self.duplicates

    def _validate_cross_workbook_duplicates(self, sheet_name: str, previous_folder: str) -> Dict[str, Any]:
        """
        Validate batch numbers against previous Daily Output workbooks.

        Args:
            sheet_name: Current worksheet name.
            previous_folder: Path to folder containing previous workbooks.

        Returns:
            Dict with 'errors', 'duplicates', and 'validation_errors'.
        """
        self.logger.info(f"Validating cross-workbook duplicates against: {previous_folder}")
        errors = []
        validation_errors = []
        duplicates: List[DuplicateRecord] = []

        try:
            checker = CrossWorkbookDuplicateChecker(previous_folder)
            indexed = checker.index_previous_workbooks()
            if not indexed:
                # Folder missing or indexing failed; log and return no errors
                self.logger.warning("Cross-workbook indexing unavailable or failed")
                return {'errors': errors, 'duplicates': duplicates}

            # Determine Batch_Number column letter
            headers = self.workbook_loader.get_headers(sheet_name)
            if not headers:
                return {'errors': errors, 'duplicates': duplicates}

            try:
                batch_idx = headers.index('Batch_Number') + 1
            except ValueError:
                self.logger.warning("Batch_Number header not found for cross-workbook check")
                return {'errors': errors, 'duplicates': duplicates}

            col_letter = get_column_letter(batch_idx)

            # Iterate current workbook batches
            batch_values = self.workbook_loader.get_column_values(sheet_name, 'Batch_Number')
            if not batch_values:
                return {'errors': errors, 'duplicates': duplicates}

            for row_num, batch_string in batch_values:
                if not batch_string:
                    continue
                batches = split_batch_numbers(batch_string)
                for batch in batches:
                    b = clean_string(batch)
                    if not b:
                        continue
                    prev_occurrences = checker.find_previous_occurrences(b)
                    if prev_occurrences:
                            # Record an error for each previous occurrence
                            for prev in prev_occurrences:
                                prev_workbook, prev_sheet, prev_row, prev_cellref = prev
                                prev_cell_ref = f"Batch_Number_{prev_row}"
                                error_msg = (f"Cross-workbook duplicate '{b}' found in previous workbook "
                                             f"{prev_workbook} sheet {prev_sheet} row {prev_row}, cell {prev_cell_ref}")
                                errors.append(error_msg)
                                validation_errors.append(ValidationError(
                                    rule_id="CROSS_WORKBOOK_DUPLICATE",
                                    error_type="DUPLICATE_CROSS_WORKBOOK",
                                    worksheet=sheet_name,
                                    row_number=row_num,
                                    column_name="Batch_Number",
                                    cell_reference=f"{col_letter}{row_num}",
                                    error_message=error_msg,
                                    invalid_value=b
                                ))
                                self.logger.warning(error_msg)

                                # Create DuplicateRecord for report
                                duplicate = DuplicateRecord(
                                    batch_number=b,
                                    worksheet=sheet_name,
                                    row_number=row_num,
                                    cell_reference=f"{col_letter}{row_num}",
                                    occurrences=len(prev_occurrences),
                                    duplicate_type="Previous Workbook"
                                )
                                duplicates.append(duplicate)

                            # Highlight duplicate in current workbook
                            try:
                                highlighter = ExcelHighlighter()
                                if highlighter.load_workbook(self.workbook_loader.current_file_path):
                                    highlighter.highlight_cell(sheet_name, row_num, col_letter, error_type="error")
                                    highlighter.save_workbook()
                                    highlighter.close()
                            except Exception:
                                # Non-fatal if highlighting fails
                                self.logger.debug("Failed to highlight cross-workbook duplicate cell")

            if not errors:
                self.logger.info("No cross-workbook duplicates found")

            return {'errors': errors, 'duplicates': duplicates, 'validation_errors': validation_errors}

        except Exception as e:
            error_msg = f"Error validating cross-workbook duplicates: {e}"
            errors.append(error_msg)
            validation_errors.append(ValidationError(
                rule_id="CROSS_WORKBOOK_VALIDATION_ERROR",
                error_type="SYSTEM_ERROR",
                worksheet=sheet_name,
                row_number=0,
                column_name="",
                cell_reference="",
                error_message=error_msg
            ))
            self.logger.error(error_msg)
            return {'errors': errors, 'duplicates': duplicates, 'validation_errors': validation_errors}

    def get_validation_errors(self) -> List[ValidationError]:
        """
        Get the list of validation errors found.
        
        Returns:
            List of ValidationError objects.
        """
        return self.validation_errors
    
    def log_validation_to_audit(self, file_path: str, result: ValidationResult) -> bool:
        """
        Log validation result to audit log.
        
        Business Rule BR014: Record workbook, date, time, validation result,
        errors, warnings and update status.
        
        Args:
            file_path: Path to the workbook that was validated.
            result: ValidationResult from validation run.
            
        Returns:
            True if audit log successful, False otherwise.
        """
        if not self.audit_manager:
            self.logger.debug("Audit manager not available")
            return False
        
        try:
            # Build audit description
            file_name = Path(file_path).name
            status = "PASSED" if result.passed else "FAILED"
            
            description = (
                f"Workbook: {file_name}\n"
                f"Status: {status}\n"
                f"Duration: {result.duration_seconds:.2f}s\n"
                f"Errors: {result.error_count}\n"
                f"Warnings: {result.warning_count}\n"
                f"Duplicates: {len(self.duplicates)}"
            )
            
            # Log to audit
            self.audit_manager.log_action(
                action="Workbook Validation",
                result=status,
                description=description
            )
            
            self.logger.info("Validation logged to audit")
            return True
            
        except Exception as e:
            self.logger.error(f"Error logging validation to audit: {e}")
            return False
