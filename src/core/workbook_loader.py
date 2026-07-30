"""
Enhanced Workbook Loader Module - Handles Excel workbook loading with openpyxl.
"""

from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Tuple, Any, Set
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from src.models.models import Workbook
from src.utils.logger import get_logger


class WorkbookLoader:
    """
    Handles loading and validating Excel workbooks.
    
    Responsible for opening workbooks, detecting worksheets,
    loading data, and providing worksheet access for validation.
    """

    # Recognized prefixes for Daily Output worksheets (case-insensitive, underscore-tolerant)
    DAILY_OUTPUT_PREFIXES = [
        "DAILY OUTPUT FILE",
        "DAILY OUTPUT",
        "DAILY_OUTPUT_FILE",
        "DAILY_OUTPUT",
    ]
    CAPITEC_SUMMARY_PREFIX = "CAPITEC SUMMARY FILE REPORT"
    
    # Required headers for Daily Output File (NEW official standard)
    REQUIRED_HEADERS = {
        'Order_no',
        'Order_Creation_Date',
        'Branch_Code',
        'Branch_Name',
        'Card_Type',
        'Number_of_Batches',
        'Waybill_Number',
        'Batch_Number',
        'BagNumber'
    }

    def __init__(self) -> None:
        """Initialize the Workbook Loader."""
        self.logger = get_logger()
        self.workbook = None
        self.current_file_path = None

    def load_workbook(self, file_path: str) -> Optional[Workbook]:
        """
        Load an Excel workbook using openpyxl.
        
        Args:
            file_path: Path to the Excel file.
            
        Returns:
            Workbook object if successful, None otherwise.
        """
        try:
            path = Path(file_path)
            
            # Verify file exists
            if not path.exists():
                self.logger.error(f"Workbook not found: {file_path}")
                return None
            
            # Verify file is Excel format
            if path.suffix.lower() not in ['.xlsx', '.xls', '.xlsm']:
                self.logger.error(f"Invalid file format: {path.suffix}")
                return None
            
            # Load with openpyxl (read-only, data_only=True to get calculated values)
            self.workbook = openpyxl.load_workbook(file_path, data_only=True)
            self.current_file_path = file_path
            
            # Create Workbook model
            workbook_model = Workbook(
                file_path=str(path.absolute()),
                file_name=path.name,
                file_size=path.stat().st_size
            )
            
            # Detect worksheets
            sheet_names = self.workbook.sheetnames
            workbook_model.worksheets = sheet_names
            workbook_model.is_valid = True
            
            self.logger.info(f"Workbook loaded successfully: {file_path}")
            self.logger.info(f"Found {len(sheet_names)} worksheets: {sheet_names}")
            
            return workbook_model
            
        except PermissionError:
            self.logger.error(f"Permission denied accessing workbook: {file_path}")
            return None
        except Exception as e:
            self.logger.error(f"Error loading workbook: {e}")
            return None

    def _normalize_sheet_name(self, name: str) -> str:
        """
        Normalize a sheet name for comparison.
        
        - Converts to uppercase
        - Replaces underscores with spaces (so "DAILY_OUTPUT" == "DAILY OUTPUT")
        - Strips leading/trailing whitespace
        
        Returns:
            Normalized string for pattern matching.
        """
        return str(name).strip().upper().replace('_', ' ')

    def _is_daily_output_sheet(self, name: str) -> bool:
        """
        Check if a worksheet name matches any recognized Daily Output pattern.
        
        Matches are:
        - Case-insensitive
        - Underscore-tolerant (underscores treated as spaces)
        - Leading/trailing whitespace ignored
        - Any recognized prefix match at the start of the normalized name
        
        Returns:
            True if the sheet appears to be a Daily Output worksheet, False otherwise.
        """
        normalized = self._normalize_sheet_name(name)
        for prefix in self.DAILY_OUTPUT_PREFIXES:
            normalized_prefix = self._normalize_sheet_name(prefix)
            if normalized.startswith(normalized_prefix):
                return True
        return False

    def _detect_date_in_sheet_name(self, sheet_name: str) -> Optional[datetime]:
        """
        Attempt to extract a date from a potential Daily Output sheet name.
        
        The date is optional — a valid Daily Output sheet may or may not have 
        a date embedded in its name. This method scans all whitespace-separated 
        tokens in the name looking for date-like patterns.
        
        Supported date formats:
        - DD-MM-YYYY, DD-MM-YY
        - DD/MM/YYYY, DD/MM/YY
        - "Month YYYY" (e.g. "July 2026", "Jul 2026")
        - DD-Month-YYYY (e.g. "15-July-2026")
        
        Returns:
            Parsed datetime if found, None otherwise.
        """
        cleaned = str(sheet_name).strip()
        tokens = cleaned.replace('_', ' ').split()
        
        for token in tokens:
            stripped = token.strip()
            # Numeric date formats
            for fmt in ("%d-%m-%Y", "%d-%m-%y", "%d/%m/%Y", "%d/%m/%y"):
                try:
                    return datetime.strptime(stripped, fmt)
                except Exception:
                    continue
            # Month name formats
            for fmt in ("%B %Y", "%b %Y", "%d-%B-%Y", "%d-%b-%Y", "%B-%Y", "%b-%Y"):
                try:
                    return datetime.strptime(stripped, fmt)
                except Exception:
                    continue
        return None

    def list_daily_output_sheets(self) -> List[Tuple[str, Optional[datetime]]]:
        """
        List all Daily Output worksheets with their detected dates.
        
        Inspects every worksheet in the workbook and classifies it as:
        - MATCHED with date: sheet matches a Daily Output pattern AND has a parseable date
        - MATCHED (no date): sheet matches a Daily Output pattern but no date found
        - REJECTED: sheet does not match any Daily Output pattern
        
        Results are sorted by date ascending (earliest first).
        Sheets without dates use a sentinel date (1900-01-01) so they appear after 
        dated sheets but are still available for processing.
        
        Returns:
            List of tuples (original_sheet_name, parsed_date_or_sentinel).
            Empty list if no workbook loaded or no matching sheets found.
        """
        sheets = []
        if not self.workbook:
            return sheets

        for name in self.workbook.sheetnames:
            cleaned = str(name).strip()
            if self._is_daily_output_sheet(cleaned):
                parsed_date = self._detect_date_in_sheet_name(cleaned)
                if parsed_date:
                    self.logger.info(f"Daily Output sheet MATCHED with date: '{name}' (parsed: {parsed_date})")
                    sheets.append((name, parsed_date))
                else:
                    self.logger.info(f"Daily Output sheet MATCHED (no date): '{name}' — included as fallback")
                    # Sentinel date so sheets-with-dates sort first, but sheet is still usable
                    sheets.append((name, datetime(1900, 1, 1)))
            else:
                self.logger.info(f"REJECTED sheet: '{name}' — does not match Daily Output pattern")

        # sort by date ascending (sheets without dates — sentinel 1900 — appear first, which is fine)
        sheets.sort(key=lambda x: x[1])
        return sheets

    def detect_daily_output_sheet(self) -> Optional[str]:
        """
        Detect the most recent (newest) Daily Output worksheet.
        
        Uses list_daily_output_sheets() to find all matching sheets,
        then returns the name of the latest one (last in sorted order).
        
        Returns:
            Sheet name if found, None otherwise.
        """
        if not self.workbook:
            self.logger.warning("No workbook loaded")
            return None

        sheets = self.list_daily_output_sheets()
        if not sheets:
            self.logger.warning("Daily Output File sheet not found in workbook")
            return None

        latest_sheet = sheets[-1][0]
        self.logger.info(f"Selected Daily Output sheet: '{latest_sheet}'")
        return latest_sheet

    def detect_capitec_summary_sheet(self) -> Optional[str]:
        """
        Detect the Capitec Summary File Report worksheet.
        
        Returns:
            Sheet name if found, None otherwise.
        """
        if not self.workbook:
            self.logger.warning("No workbook loaded")
            return None
        
        for sheet_name in self.workbook.sheetnames:
            if sheet_name.startswith(self.CAPITEC_SUMMARY_PREFIX):
                self.logger.info(f"Found Capitec Summary sheet: '{sheet_name}'")
                return sheet_name
        
        self.logger.warning("Capitec Summary File Report sheet not found")
        return None

    def get_worksheet(self, sheet_name: str) -> Optional[Worksheet]:
        """
        Get a worksheet by name.
        
        Args:
            sheet_name: Name of the worksheet.
            
        Returns:
            Worksheet object if found, None otherwise.
        """
        if not self.workbook:
            self.logger.warning("No workbook loaded")
            return None
        
        try:
            return self.workbook[sheet_name]
        except KeyError:
            self.logger.error(f"Worksheet not found: {sheet_name}")
            return None

    def _normalize_header(self, value: Optional[Any]) -> str:
        if value is None:
            return ''
        return str(value).strip().lower().replace(' ', '_')

    def _find_header_row(self, worksheet: Worksheet) -> Tuple[int, List[str]]:
        """
        Find the header row in a worksheet.
        
        Scans rows 1-20 looking for a row that contains all required headers.
        
        Returns:
            Tuple of (header_row_index, list_of_header_values).
        """
        # Precompute the set of normalized required headers
        required_normalized = {
            self._normalize_header(h)
            for h in self.REQUIRED_HEADERS
        }

        for row_idx, row in enumerate(worksheet.iter_rows(max_row=20, values_only=True), start=1):
            raw_headers = [str(cell).strip() if cell is not None else '' for cell in row]
            
            # Build a set of normalized header names found in this row
            found_normalized = set()
            for h in raw_headers:
                normalized = self._normalize_header(h)
                if normalized in required_normalized:
                    found_normalized.add(normalized)
            
            if required_normalized.issubset(found_normalized):
                if row_idx != 1:
                    self.logger.info(
                        f"Detected header row {row_idx} in worksheet {worksheet.title} "
                        f"(headers: {raw_headers})"
                    )
                return row_idx, raw_headers

        self.logger.warning(
            f"Could not detect required headers in first 20 rows of {worksheet.title}; "
            f"falling back to row 1"
        )
        headers = [str(cell.value).strip() if cell.value is not None else '' for cell in worksheet[1]]
        return 1, headers

    def get_headers(self, sheet_name: str) -> Optional[List[str]]:
        """
        Get headers from a worksheet.
        
        Args:
            sheet_name: Name of the worksheet.
            
        Returns:
            List of header names if successful, None otherwise.
        """
        worksheet = self.get_worksheet(sheet_name)
        if not worksheet:
            return None
        
        try:
            _, raw_headers = self._find_header_row(worksheet)
            
            self.logger.debug(f"Headers from {sheet_name}: {raw_headers}")
            return raw_headers
            
        except Exception as e:
            self.logger.error(f"Error getting headers from {sheet_name}: {e}")
            return None

    def get_data_rows(self, sheet_name: str) -> Optional[List[Dict[str, Any]]]:
        """
        Get all data rows from a worksheet as dictionaries with header keys.
        
        Args:
            sheet_name: Name of the worksheet.
            
        Returns:
            List of dictionaries with row data (keys are header names),
            None on error.
        """
        worksheet = self.get_worksheet(sheet_name)
        if not worksheet:
            return None
        
        try:
            all_values = list(worksheet.iter_rows(min_row=1, values_only=True))
            total_raw_rows = len(all_values)
            non_empty_rows = sum(1 for row in all_values if any(cell is not None and str(cell).strip() != '' for cell in row))
            self.logger.info(f"Sheet {sheet_name} max_row={worksheet.max_row}, total_raw_rows={total_raw_rows}, non_empty_rows={non_empty_rows}")

            header_row_idx, headers = self._find_header_row(worksheet)
            if not headers:
                return None
            
            self.logger.info(f"Loading data rows from sheet: {sheet_name}")
            self.logger.info(f"Using header row: {header_row_idx}")
            self.logger.info(f"Headers: {headers}")

            rows = []
            raw_rows = []
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=header_row_idx + 1, values_only=False), start=header_row_idx + 1):
                raw_rows.append(row)
                row_data = {}
                for col_idx, (header, cell) in enumerate(zip(headers, row), start=1):
                    # Get cell value and reference
                    value = cell.value
                    cell_ref = cell.coordinate
                    
                    row_data[header] = {
                        'value': value,
                        'cell_ref': cell_ref,
                        'row': row_idx,
                        'col': col_idx
                    }
                
                rows.append(row_data)

            self.logger.info(f"Total raw rows read from Excel: {len(raw_rows)}")
            self.logger.info("First 5 rows of extracted data: %s", rows[:5])
            self.logger.info(f"Loaded {len(rows)} data rows from {sheet_name}")
            return rows
            
        except Exception as e:
            self.logger.error(f"Error getting data rows from {sheet_name}: {e}")
            return None

    def get_column_values(self, sheet_name: str, column_name: str) -> Optional[List[Tuple[int, str]]]:
        """
        Get all values from a specific column.
        
        Args:
            sheet_name: Name of the worksheet.
            column_name: Name of the column.
            
        Returns:
            List of (row_number, value) tuples, None on error.
        """
        worksheet = self.get_worksheet(sheet_name)
        if not worksheet:
            return None
        
        try:
            headers = self.get_headers(sheet_name)
            if not headers:
                return None
            
            # Find the column index
            col_normalized = self._normalize_header(column_name)
            target_col = None
            for h in headers:
                if self._normalize_header(h) == col_normalized:
                    target_col = h
                    break
            
            if not target_col:
                self.logger.warning(
                    f"Column '{column_name}' not found in {sheet_name}"
                )
                return None
            
            col_idx = headers.index(target_col) + 1
            values = []
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_col=col_idx), start=2):
                cell = row[col_idx - 1]
                value = cell.value
                if value is not None:
                    values.append((row_idx, str(value).strip()))
            
            self.logger.debug(f"Retrieved {len(values)} values from column '{column_name}'")
            return values
            
        except Exception as e:
            self.logger.error(f"Error getting column values: {e}")
            return None

    def close(self) -> None:
        """Close the workbook."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.logger.info("Workbook closed")