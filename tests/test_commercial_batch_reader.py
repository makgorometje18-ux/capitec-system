import openpyxl
from datetime import datetime

from src.core.commercial_batch_reader import read_commercial_batches
from src.models.models import CommercialBatch


def create_commercial_batch_sheet(workbook: openpyxl.Workbook, rows: list) -> None:
    worksheet = workbook.create_sheet('COMMERCIAL BATCH')
    headers = ['Batch Number', 'Card Type', 'Quantity', 'Order Number']
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)


def test_read_commercial_batches_normal_rows(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    create_commercial_batch_sheet(
        wb,
        [
            ['BATCH001', 'SIM', 100, 'ORD001'],
            ['BATCH002', 'DMCCLS', 200, 'ORD002'],
        ]
    )

    path = tmp_path / 'commercial_batches.xlsx'
    wb.save(path)

    batches = read_commercial_batches(openpyxl.load_workbook(path))

    assert len(batches) == 2
    assert batches[0] == CommercialBatch(
        batch_number='BATCH001',
        card_type='SIM',
        quantity=100,
        order_number='ORD001',
        raw_row_index=2,
    )
    assert batches[1].batch_number == 'BATCH002'
    assert batches[1].card_type == 'DMCCLS'
    assert batches[1].quantity == 200
    assert batches[1].order_number == 'ORD002'
    assert batches[1].raw_row_index == 3


def test_read_commercial_batches_skips_empty_rows(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    create_commercial_batch_sheet(
        wb,
        [
            ['BATCH001', 'SIM', 100, 'ORD001'],
            [None, None, None, None],
            ['BATCH002', 'DMCCLS', 200, 'ORD002'],
        ]
    )

    path = tmp_path / 'commercial_batches_empty.xlsx'
    wb.save(path)

    batches = read_commercial_batches(openpyxl.load_workbook(path))

    assert len(batches) == 2
    assert batches[0].raw_row_index == 2
    assert batches[1].raw_row_index == 4


def test_read_commercial_batches_handles_mixed_data_types(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    create_commercial_batch_sheet(
        wb,
        [
            ['BATCH001', 'SIM', '100', 'ORD001'],
            ['BATCH002', 'DMCCLS', 200.0, None],
            ['BATCH003', 'SIM', None, 'ORD003'],
        ]
    )

    path = tmp_path / 'commercial_batches_mixed.xlsx'
    wb.save(path)

    batches = read_commercial_batches(openpyxl.load_workbook(path))

    assert len(batches) == 3
    assert batches[0].quantity == 100
    assert batches[0].order_number == 'ORD001'
    assert batches[1].quantity == 200
    assert batches[1].order_number is None
    assert batches[2].quantity == 0
    assert batches[2].order_number == 'ORD003'
