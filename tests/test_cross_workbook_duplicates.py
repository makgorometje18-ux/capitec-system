"""
Tests for Cross-Workbook Duplicate Detection.
"""
import shutil
from pathlib import Path
import openpyxl
import pytest

from src.core.validation_engine import ValidationEngine
from src.core.sample_workbook_generator import generate_sample_workbooks


SAMPLE_DIR = Path("sample_files")
PREV_DIR = SAMPLE_DIR / "previous"


def create_previous_workbook(path: Path, batches: list, sheet_name: str = None):
    wb = openpyxl.Workbook()
    sheet_name = sheet_name or "DAILY OUTPUT FILE 01-07-2026"
    ws = wb.active
    ws.title = sheet_name
    # Create simple headers with Batch_No in column H (but header lookup uses name only)
    headers = ["Order_No", "Order_Creation_Date", "Branch_Code", "Branch_Name", "Card_Type", "No_of_Batches", "Waybill_No", "Batch_No", "Bag_No"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx).value = h
    
    for i, batch in enumerate(batches, start=2):
        # Put batch in Batch_No column (8th column) and align with the new header set
        ws.cell(row=i, column=1).value = i - 1
        ws.cell(row=i, column=2).value = 2026
        ws.cell(row=i, column=3).value = 4000 + i
        ws.cell(row=i, column=4).value = f"Branch {i}"
        ws.cell(row=i, column=5).value = "SIM"
        ws.cell(row=i, column=6).value = 1
        ws.cell(row=i, column=7).value = f"WB{i}"
        ws.cell(row=i, column=8).value = batch
        ws.cell(row=i, column=9).value = "'00000001"

    wb.save(path)
    wb.close()


@pytest.fixture(autouse=True)
def ensure_samples_exist(tmp_path_factory):
    # Ensure sample workbooks generated
    generate_sample_workbooks()
    # Ensure previous folder exists and is clean
    if PREV_DIR.exists():
        shutil.rmtree(PREV_DIR)
    PREV_DIR.mkdir(parents=True, exist_ok=True)
    yield
    # Cleanup
    if PREV_DIR.exists():
        shutil.rmtree(PREV_DIR)


def test_no_duplicates_found():
    # Create previous workbook with different batches
    prev_file = PREV_DIR / "prev1.xlsx"
    create_previous_workbook(prev_file, ["50001", "50002"])    

    engine = ValidationEngine()
    result = engine.validate_complete_workbook("sample_files/sample_invalid.xlsx", previous_folder=str(PREV_DIR))

    cross_dups = [d for d in engine.get_duplicates() if d.duplicate_type == "Previous Workbook"]
    assert len(cross_dups) == 0
    assert result.passed == False  # overall validation still fails due to current workbook errors


def test_one_duplicate_found():
    # Create previous workbook containing one batch that matches current sample_invalid (30001 exists)
    prev_file = PREV_DIR / "prev2.xlsx"
    create_previous_workbook(prev_file, ["30001", "70002"])    

    engine = ValidationEngine()
    result = engine.validate_complete_workbook("sample_files/sample_invalid.xlsx", previous_folder=str(PREV_DIR))

    cross_dups = [d for d in engine.get_duplicates() if d.duplicate_type == "Previous Workbook"]
    assert len(cross_dups) >= 1
    # Ensure error messages contain 'Cross-workbook duplicate'
    cw_errors = [e for e in result.errors if 'Cross-workbook duplicate' in e]
    assert len(cw_errors) >= 1


def test_multiple_duplicates_found():
    # Create multiple previous workbooks with overlapping batches
    prev_file1 = PREV_DIR / "prev3.xlsx"
    prev_file2 = PREV_DIR / "prev4.xlsx"
    create_previous_workbook(prev_file1, ["20001", "90002"])    
    create_previous_workbook(prev_file2, ["30001", "20001"])    

    engine = ValidationEngine()
    result = engine.validate_complete_workbook("sample_files/sample_invalid.xlsx", previous_folder=str(PREV_DIR))

    cross_dups = [d for d in engine.get_duplicates() if d.duplicate_type == "Previous Workbook"]
    # We expect at least two previous occurrences recorded
    assert len(cross_dups) >= 2


def test_missing_previous_folder():
    # Point to a missing folder
    missing = PREV_DIR / "does_not_exist"
    if missing.exists():
        shutil.rmtree(missing)

    engine = ValidationEngine()
    # Should not raise
    result = engine.validate_complete_workbook("sample_files/sample_invalid.xlsx", previous_folder=str(missing))
    # No cross-workbook duplicates expected
    cross_dups = [d for d in engine.get_duplicates() if d.duplicate_type == "Previous Workbook"]
    assert len(cross_dups) == 0


def test_empty_previous_folder():
    # PREV_DIR exists but empty (fixture ensures it's empty)
    engine = ValidationEngine()
    result = engine.validate_complete_workbook("sample_files/sample_invalid.xlsx", previous_folder=str(PREV_DIR))
    cross_dups = [d for d in engine.get_duplicates() if d.duplicate_type == "Previous Workbook"]
    assert len(cross_dups) == 0

 