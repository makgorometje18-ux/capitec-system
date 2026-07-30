import openpyxl
from datetime import datetime
from pathlib import Path

from src.core.summary_reconciliation_engine import SummaryReconciliationEngine
from src.gui.dashboard import build_summary_update_preview_text


def create_daily_sheet(workbook: openpyxl.Workbook, title: str, rows: list) -> None:
    worksheet = workbook.create_sheet(title)
    headers = ['Order_no', 'Order_Creation_Date', 'Branch_Code', 'Branch_Name', 'Card_Type', 'Number_of_Batches', 'Waybill_Number', 'Batch_Number', 'BagNumber']
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)


def create_summary_sheet(workbook: openpyxl.Workbook, rows: list) -> None:
    worksheet = workbook.create_sheet('CAPITEC SUMMARY FILE REPORT')
    headers = ['Files received', 'Total quantity received', 'Quantity in stock', 'Quantity dispatched', 'Comment']
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)


def test_summary_update_preview_displays_row_updates_only(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    create_daily_sheet(
        wb,
        'DAILY OUTPUT FILE 03-07-2026',
        [
            [1, datetime(2026, 7, 3).date(), 'BR003', 'Branch C', 'SIM', 2, 'WB003', '30001|30002', "'00034567"],
            [2, datetime(2026, 7, 3).date(), 'BR004', 'Branch D', 'DMCCLS', 1, 'WB004', '40001', "'00045678"],
        ]
    )
    create_summary_sheet(
        wb,
        [
            ['P_001_BANK', 1000, 500, 200, 'Bank card file'],
            ['C-Connect batch 1', 800, 300, 100, 'SIM file batch'],
            ['SIM Orders', 2, None, None, 'Calculated summary metric'],
            ['DMCCLS Orders', 1, None, None, 'Calculated summary metric'],
        ]
    )

    path = tmp_path / 'preview_integrated.xlsx'
    wb.save(path)

    engine = SummaryReconciliationEngine()
    analysis = engine.analyze(str(path))

    preview_text = build_summary_update_preview_text(analysis)

    assert 'Files received' in preview_text
    assert 'Current Quantity In Stock' in preview_text
    assert 'New Quantity In Stock' in preview_text
    assert 'Current Quantity Dispatched' in preview_text
    assert 'New Quantity Dispatched' in preview_text
    assert 'Card Change' in preview_text

    assert 'P_001_BANK' in preview_text
    assert 'C-Connect batch 1' in preview_text

    assert 'Row Type' not in preview_text
    assert 'Changes Required' not in preview_text
    assert 'SIM Orders' not in preview_text
    assert 'DMCCLS Orders' not in preview_text
    assert 'Total Orders' not in preview_text
    assert 'N/A' not in preview_text
    assert '500' in preview_text
    assert '200' in preview_text


def test_summary_update_preview_parses_formula_cells_and_alias_headers(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    create_daily_sheet(
        wb,
        'DAILY OUTPUT FILE 03-07-2026',
        [
            [1, datetime(2026, 7, 3).date(), 'BR003', 'Branch C', 'SIM', 2, 'WB003', '30001|30002', "'00034567"],
            [2, datetime(2026, 7, 3).date(), 'BR004', 'Branch D', 'DMCCLS', 2, 'WB004', '40001|40002', "'00045678"],
        ]
    )
    worksheet = wb.create_sheet('CAPITEC SUMMARY FILE REPORT')
    worksheet.append(['Files received', 'Total quantity received', 'Current Quantity In Stock', 'Current Quantity Dispatched', 'Comment'])
    worksheet.append(['P_001_BANK', 1000, '=500', '=200', 'Bank card formula values'])
    worksheet.append(['C-Connect batch 1', 800, '=300', '=100', 'SIM formula values'])

    path = tmp_path / 'preview_formula_alias.xlsx'
    wb.save(path)

    engine = SummaryReconciliationEngine()
    analysis = engine.analyze(str(path))

    preview_text = build_summary_update_preview_text(analysis)

    assert 'P_001_BANK' in preview_text
    assert 'C-Connect batch 1' in preview_text
    assert '500' in preview_text
    assert '200' in preview_text
    assert '300' in preview_text
    assert '100' in preview_text
    assert 'N/A' not in preview_text
