import openpyxl
from pathlib import Path

from src.core.summary_reconciliation_engine import SummaryReconciliationEngine
from src.core.summary_updater import SummaryWorksheetUpdater
from src.core.workbook_loader import WorkbookLoader


def test_summary_analysis_to_updater_integration(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet('CAPITEC SUMMARY FILE REPORT')
    headers = ['Files received', 'Total quantity received', 'Quantity in stock', 'Quantity dispatched', 'Comment']
    ws_summary.append(headers)
    ws_summary.append(['SIM Orders', 1, None, None, ''])
    ws_summary.append(['SIM Cards', 400, None, None, ''])
    ws_summary.append(['DMCCLS Orders', 0, None, None, ''])
    ws_summary.append(['DMCCLS Cards', 0, None, None, ''])
    ws_summary.append(['Total Orders', 1, None, None, ''])
    ws_summary.append(['Total Cards', 200, None, None, ''])

    ws_daily = wb.create_sheet('DAILY OUTPUT FILE 03-07-2026')
    daily_headers = ['Order_No', 'Order_Creation_Date', 'Branch_Code', 'Branch_Name', 'Card_Type', 'No_of_Batches', 'Waybill_No', 'Batch_No', 'Bag_No']
    ws_daily.append(daily_headers)
    ws_daily.append([1, '2026-07-03', 'BR003', 'Branch C', 'SIM', 1, 'WB003', '30001|30002|30003|30004|30005', "'00034567"])

    path = tmp_path / 'integration_summary.xlsx'
    wb.save(path)

    engine = SummaryReconciliationEngine()
    analysis = engine.analyze(str(path))
    assert analysis.preview_ready is True

    loader = engine.loader
    updater = SummaryWorksheetUpdater()
    result = updater.update_summary_worksheet(analysis, loader)

    assert result.success is True
    assert len(result.updated_fields) > 0
    assert len(result.failed_fields) == 0
    assert result.elapsed_time >= 0
    assert loader.workbook is not None
    assert loader.current_file_path == str(path)
