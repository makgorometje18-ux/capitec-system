import shutil
from pathlib import Path
import openpyxl
import pytest

from src.core.reconciliation_engine import ReconciliationEngine
from src.core.sample_workbook_generator import create_valid_daily_output_workbook, generate_sample_workbooks


def test_newest_sheet_detection(tmp_path):
    # create workbook with multiple daily sheets
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "DAILY OUTPUT FILE 01-07-2026"
    wb.create_sheet("DAILY OUTPUT FILE 02-07-2026")
    wb.create_sheet("DAILY OUTPUT FILE 03-07-2026")
    p = tmp_path / "multi_daily.xlsx"
    wb.save(p)

    engine = ReconciliationEngine()
    assert engine.load_workbook(str(p))
    sheets = engine.list_daily_output_sheets()
    assert len(sheets) == 3
    active = engine.select_active_sheet()
    assert active == "DAILY OUTPUT FILE 03-07-2026"


def test_sim_calculation(tmp_path):
    # Generate sample files and use valid sample
    generate_sample_workbooks()
    sample = Path("sample_files") / "sample_valid.xlsx"
    engine = ReconciliationEngine()
    assert engine.load_workbook(str(sample))
    active = engine.select_active_sheet()
    stats = engine.compute_card_statistics(active)
    # From sample_valid: SIM orders = 2 + 1 = 3; SIM cards = 3 * 200 = 600
    assert stats.sim_orders == 3
    assert stats.sim_cards == 600


def test_negative_stock_prevention(tmp_path):
    # Create a workbook with Capitec Summary and a waybill row with low stock
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "CAPITEC SUMMARY FILE REPORT"
    headers = ['Files received', 'Total quantity received', 'Quantity in stock', 'Quantity dispatched', 'Comment']
    ws_summary.append(headers)
    # Add a waybill row with stock 100
    ws_summary.append(['WBNEG', 100, 100, 0, ''])

    # Add a daily sheet with a DMCCLS entry that requests more than stock
    ws_daily = wb.create_sheet('DAILY OUTPUT FILE 03-07-2026')
    headers2 = ['Order_no', 'Order_Creation_Date', 'Branch_Code', 'Branch_Name', 'Card_Type', 'Number_of_Batches', 'Waybill_Number', 'Batch_Number', 'BagNumber']
    ws_daily.append(headers2)
    ws_daily.append([1, '2026-07-03', 'BRX', 'Branch X', 'DMCCLS', 1, 'WBNEG', '90001', "'00000123"])

    p = tmp_path / 'neg_stock.xlsx'
    wb.save(p)

    engine = ReconciliationEngine()
    assert engine.load_workbook(str(p))
    active = engine.select_active_sheet()
    stats = engine.compute_card_statistics(active)
    ok, msgs = engine.update_summary_report(active, stats)
    assert ok is False
    assert any('Negative stock' in m or 'Not enough' in m for m in msgs)


def test_dynamic_cconnect_batches(tmp_path):
    """Test that dynamic C-Connect batch lookup works for any number of batches."""
    # Create a workbook with multiple C-Connect batches (1, 2, 3)
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "CAPITEC SUMMARY FILE REPORT"
    headers = ['Files received', 'Total quantity received', 'Quantity in stock', 'Quantity dispatched', 'Comment']
    ws_summary.append(headers)
    # Add 3 C-Connect batches with stock
    ws_summary.append(['C-Connect batch 1', 100, 1000, 0, ''])
    ws_summary.append(['C-Connect batch 2', 100, 500, 0, ''])
    ws_summary.append(['C-Connect batch 3', 100, 300, 0, ''])
    # Add a non-C-Connect row that should be ignored
    ws_summary.append(['Other Item', 50, 200, 0, ''])

    # Add a daily sheet with 1000 SIM cards to dispatch
    ws_daily = wb.create_sheet('DAILY OUTPUT FILE 03-07-2026')
    headers2 = ['Order_no', 'Order_Creation_Date', 'Branch_Code', 'Branch_Name', 'Card_Type', 'Number_of_Batches', 'Waybill_Number', 'Batch_Number', 'BagNumber']
    ws_daily.append(headers2)
    # 5 SIM orders * 200 cards/order = 1000 cards
    ws_daily.append([1, '2026-07-03', 'BRX', 'Branch X', 'SIM', 5, 'WBX', '11001|11002|11003|11004|11005', "'00000123"])

    p = tmp_path / 'multi_cconnect.xlsx'
    wb.save(p)

    engine = ReconciliationEngine()
    assert engine.load_workbook(str(p))
    active = engine.select_active_sheet()
    stats = engine.compute_card_statistics(active)
    assert stats.sim_cards == 1000
    ok, msgs = engine.update_summary_report(active, stats)
    assert ok is True
    # Verify all 3 C-Connect batches were updated
    assert any('C-Connect batch 1' in m for m in msgs)
    assert any('C-Connect batch 2' in m for m in msgs)
    assert any('C-Connect batch 3' in m for m in msgs)
    # Verify the Other Item row was NOT updated
    assert not any('Other Item' in m for m in msgs)
