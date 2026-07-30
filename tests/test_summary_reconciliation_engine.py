import openpyxl
import pytest
from datetime import datetime
from pathlib import Path

from src.core.summary_reconciliation_engine import SummaryReconciliationEngine
from src.gui.dashboard import build_summary_update_preview_text
from src.models.models import SummaryAnalysis


def create_daily_sheet(workbook: openpyxl.Workbook, title: str, rows: list) -> None:
    worksheet = workbook.create_sheet(title)
    headers = ['Order_no', 'Order_Creation_Date', 'Branch_Code', 'Branch_Name', 'Card_Type', 'Number_of_Batches', 'Waybill_Number', 'Batch_Number', 'BagNumber']
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)


def create_summary_sheet(workbook: openpyxl.Workbook, rows: list) -> None:
    worksheet = workbook.create_sheet('CAPITEC SUMMARY FILE REPORT')
    headers = ['Files received', 'Total quantity received', 'Quantity in stock', 'Quantity dispatched', 'Comment']
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)


def test_summary_analysis_selects_latest_daily_sheet(tmp_path):
    wb = openpyxl.Workbook()
    # Keep the default sheet if it's blank and irrelevant
    wb.remove(wb.active)

    create_daily_sheet(
        wb,
        'DAILY OUTPUT FILE 01-07-2026',
        [
            [1, datetime(2026, 7, 1).date(), 'BR001', 'Branch A', 'SIM', 2, 'WB001', '10001|10002', "'00012345"],
        ]
    )
    create_daily_sheet(
        wb,
        'DAILY OUTPUT FILE 03-07-2026',
        [
            [1, datetime(2026, 7, 3).date(), 'BR002', 'Branch B', 'DMCCLS', 1, 'WB002', '20001', "'00023456"],
            [2, datetime(2026, 7, 3).date(), 'BR003', 'Branch C', 'SIM', 1, 'WB003', '30001', "'00034567"],
        ]
    )
    create_daily_sheet(
        wb,
        'DAILY OUTPUT FILE 02-07-2026',
        [
            [1, datetime(2026, 7, 2).date(), 'BR004', 'Branch D', 'SIM', 3, 'WB004', '40001|40002|40003', "'00045678"],
        ]
    )

    create_summary_sheet(
        wb,
        [
            ['C-Connect batch 1', 100, 80, 20, ''],
            ['C-Connect batch 2', 200, 180, 20, ''],
            ['SIM Orders', 3, None, None, ''],
            ['DMCCLS Orders', 1, None, None, ''],
        ]
    )

    path = tmp_path / 'summary_test.xlsx'
    wb.save(path)

    engine = SummaryReconciliationEngine()
    analysis = engine.analyze(str(path))

    assert analysis.summary_worksheet_name == 'CAPITEC SUMMARY FILE REPORT'
    assert analysis.latest_daily_worksheet_name == 'DAILY OUTPUT FILE 03-07-2026'
    assert analysis.sim_orders == 1
    assert analysis.sim_cards == 200
    assert analysis.dmcc_orders == 1
    assert analysis.dmcc_cards == 300
    assert analysis.total_orders == 2
    assert analysis.total_cards == 500
    assert len(analysis.cconnect_rows) == 2
    assert analysis.existing_summary_values['C-Connect batch 1']['Total quantity received'] == 100
    assert any(diff.metric_name == 'SIM Orders' for diff in analysis.differences)
    assert analysis.validation_status is True


def test_summary_analysis_handles_missing_summary_sheet(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    create_daily_sheet(
        wb,
        'DAILY OUTPUT FILE 03-07-2026',
        [
            [1, datetime(2026, 7, 3).date(), 'BR002', 'Branch B', 'DMCCLS', 1, 'WB002', '20001', "'00023456"],
        ]
    )
    path = tmp_path / 'missing_summary.xlsx'
    wb.save(path)

    engine = SummaryReconciliationEngine()
    analysis = engine.analyze(str(path))

    assert analysis.summary_worksheet_name is None
    assert analysis.latest_daily_worksheet_name is None
    assert analysis.validation_status is False


def test_summary_analysis_reads_existing_summary_values(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    create_daily_sheet(
        wb,
        'DAILY OUTPUT FILE 03-07-2026',
        [
            [1, datetime(2026, 7, 3).date(), 'BR003', 'Branch C', 'SIM', 2, 'WB003', '30001|30002', "'00034567"],
        ]
    )
    create_summary_sheet(
        wb,
        [
            ['SIM Orders', 2, None, None, 'Calculated from daily sheet'],
            ['SIM Cards', 400, None, None, ''],
        ]
    )
    path = tmp_path / 'existing_values.xlsx'
    wb.save(path)

    engine = SummaryReconciliationEngine()
    analysis = engine.analyze(str(path))

    assert analysis.existing_summary_values['SIM Orders']['Files received'] == 'SIM Orders'
    assert any(diff.metric_name == 'SIM Orders' and diff.status != 'Unknown' for diff in analysis.differences)


def test_summary_analysis_finds_summary_headers_below_title_row(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws_summary = wb.create_sheet('CAPITEC SUMMARY FILE REPORT')
    ws_summary.append(['CAPITEC SUMMARY FILE REPORT'])
    ws_summary.append(['Files received', 'Total quantity received', 'Quantity in stock', 'Quantity dispatched', 'Comment'])
    ws_summary.append(['SIM Orders', 2, None, None, 'Existing summary value'])
    ws_summary.append(['SIM Cards', 400, None, None, 'Existing summary value'])

    create_daily_sheet(
        wb,
        'DAILY OUTPUT FILE 03-07-2026',
        [
            [1, datetime(2026, 7, 3).date(), 'BR003', 'Branch C', 'SIM', 2, 'WB003', '30001|30002', "'00034567"],
        ]
    )

    path = tmp_path / 'summary_with_title.xlsx'
    wb.save(path)

    engine = SummaryReconciliationEngine()
    analysis = engine.analyze(str(path))

    assert analysis.summary_worksheet_name == 'CAPITEC SUMMARY FILE REPORT'
    assert analysis.existing_summary_values['SIM Orders']['Total quantity received'] == 2
    assert analysis.current_summary_values['SIM Orders'] == 2
    sim_orders = next(diff for diff in analysis.differences if diff.metric_name == 'SIM Orders')
    assert sim_orders.existing_value == 2
    assert sim_orders.difference == 0
    assert sim_orders.status == 'Match'


def test_summary_analysis_models_summary_rows_and_calculates_row_updates(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    create_daily_sheet(
        wb,
        'DAILY OUTPUT FILE 03-07-2026',
        [
            [1, datetime(2026, 7, 3).date(), 'BR003', 'Branch C', 'SIM', 3, 'WB003', '30001|30002|30003', "'00034567"],
            [2, datetime(2026, 7, 3).date(), 'BR004', 'Branch D', 'DMCCLS', 2, 'WB004', '40001|40002', "'00045678"],
        ]
    )
    worksheet = wb.create_sheet('CAPITEC SUMMARY FILE REPORT')
    worksheet.append(['Files received', 'Total quantity received', 'Quantity in stock', 'Quantity dispatched', 'Comment'])
    worksheet.append(['P_001_BANK', 1000, 500, 200, 'Bank card file'])
    worksheet.append(['C-Connect batch 1', 800, 300, 100, 'SIM file batch'])

    path = tmp_path / 'summary_row_model.xlsx'
    wb.save(path)

    engine = SummaryReconciliationEngine()
    analysis = engine.analyze(str(path))

    assert analysis.summary_rows
    assert len(analysis.files_received_rows) == 2
    assert len(analysis.bank_card_rows) == 1
    assert len(analysis.sim_rows) == 1

    bank_row = analysis.bank_card_rows[0]
    assert bank_row.row_type == 'BANK_CARD'
    assert bank_row.calculated_card_change == analysis.dmcc_cards
    assert bank_row.new_quantity_in_stock == 500 - analysis.dmcc_cards
    assert bank_row.new_quantity_dispatched == 200 + analysis.dmcc_cards
    assert bank_row.changes_required is True

    sim_row = analysis.sim_rows[0]
    assert sim_row.row_type == 'SIM'
    assert sim_row.calculated_card_change == analysis.sim_cards
    assert sim_row.new_quantity_in_stock == 300 - analysis.sim_cards
    assert sim_row.new_quantity_dispatched == 100 + analysis.sim_cards
    assert sim_row.changes_required is True


def test_summary_analysis_parses_formula_cells_in_summary_sheet(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    create_daily_sheet(
        wb,
        'DAILY OUTPUT FILE 03-07-2026',
        [
            [1, datetime(2026, 7, 3).date(), 'BR003', 'Branch C', 'SIM', 2, 'WB003', '30001|30002', "'00034567"],
            [2, datetime(2026, 7, 3).date(), 'BR004', 'Branch D', 'DMCCLS', 2, 'WB004', '40001|40002', "'00045678"],
        ]
    )
    worksheet = wb.create_sheet('CAPITEC SUMMARY FILE REPORT')
    worksheet.append(['Files received', 'Total quantity received', 'Current Quantity In Stock', 'Current Quantity Dispatched', 'Comment'])
    worksheet.append(['P_001_BANK', 1000, '=500', '=200', 'Bank card formula values'])
    worksheet.append(['C-Connect batch 1', 800, '=300', '=100', 'SIM formula values'])

    path = tmp_path / 'summary_formula_cells.xlsx'
    wb.save(path)

    engine = SummaryReconciliationEngine()
    analysis = engine.analyze(str(path))

    assert len(analysis.summary_rows) == 2
    bank_row = next(row for row in analysis.summary_rows if row.row_type == 'BANK_CARD')
    sim_row = next(row for row in analysis.summary_rows if row.row_type == 'SIM')

    assert bank_row.quantity_in_stock == 500
    assert bank_row.quantity_dispatched == 200
    assert sim_row.quantity_in_stock == 300
    assert sim_row.quantity_dispatched == 100
    assert bank_row.changes_required is True
    assert sim_row.changes_required is True

    preview = build_summary_update_preview_text(analysis)
    assert 'P_001_BANK' in preview
    assert 'C-Connect batch 1' in preview
    assert '500' in preview
    assert '200' in preview
    assert '300' in preview
    assert '100' in preview


def test_summary_analysis_selects_active_p_row_and_ignores_history(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    create_daily_sheet(
        wb,
        'DAILY OUTPUT FILE 03-07-2026',
        [
            [1, datetime(2026, 7, 3).date(), 'BR003', 'Branch C', 'SIM', 2, 'WB003', '30001|30002', "'00034567"],
            [2, datetime(2026, 7, 3).date(), 'BR004', 'Branch D', 'DMCCLS', 2, 'WB004', '40001|40002', "'00045678"],
        ]
    )
    worksheet = wb.create_sheet('CAPITEC SUMMARY FILE REPORT')
    worksheet.append(['Files received', 'Total quantity received', 'Quantity in stock', 'Quantity dispatched', 'Comment'])
    worksheet.append(['P_OLD_HISTORY', 1000, 100, 0, 'Historical bank record'])
    worksheet.append(['P_001_CURRENT', 1000, 500, 200, 'Active bank record'])
    worksheet.append(['P_002_NEXT', 1000, 100, 50, 'Later bank record'])

    path = tmp_path / 'active_bank_inventory.xlsx'
    wb.save(path)

    engine = SummaryReconciliationEngine()
    analysis = engine.analyze(str(path))

    active_bank_row = next(row for row in analysis.summary_rows if row.item_name == 'P_001_CURRENT')
    historical_bank_row = next(row for row in analysis.summary_rows if row.item_name == 'P_OLD_HISTORY')
    later_bank_row = next(row for row in analysis.summary_rows if row.item_name == 'P_002_NEXT')

    assert active_bank_row.changes_required is True
    assert active_bank_row.new_quantity_in_stock == 500 - analysis.dmcc_cards
    assert active_bank_row.new_quantity_dispatched == 200 + analysis.dmcc_cards
    assert historical_bank_row.changes_required is False
    assert later_bank_row.changes_required is False


def test_summary_analysis_selects_active_c_connect_row_and_ignores_history(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    create_daily_sheet(
        wb,
        'DAILY OUTPUT FILE 03-07-2026',
        [
            [1, datetime(2026, 7, 3).date(), 'BR003', 'Branch C', 'SIM', 2, 'WB003', '30001|30002', "'00034567"],
            [2, datetime(2026, 7, 3).date(), 'BR004', 'Branch D', 'DMCCLS', 2, 'WB004', '40001|40002', "'00045678"],
        ]
    )
    worksheet = wb.create_sheet('CAPITEC SUMMARY FILE REPORT')
    worksheet.append(['Files received', 'Total quantity received', 'Quantity in stock', 'Quantity dispatched', 'Comment'])
    worksheet.append(['C-Connect batch 1', 800, 500, 100, 'Active SIM record'])
    worksheet.append(['C-Connect batch 2', 800, 0, 50, 'Historical SIM record'])
    worksheet.append(['C-Connect batch 3', 800, 100, 0, 'Next SIM record'])

    path = tmp_path / 'active_sim_inventory.xlsx'
    wb.save(path)

    engine = SummaryReconciliationEngine()
    analysis = engine.analyze(str(path))

    active_sim_row = next(row for row in analysis.summary_rows if row.item_name == 'C-Connect batch 1')
    historical_sim_row = next(row for row in analysis.summary_rows if row.item_name == 'C-Connect batch 2')
    next_sim_row = next(row for row in analysis.summary_rows if row.item_name == 'C-Connect batch 3')

    assert active_sim_row.changes_required is True
    assert active_sim_row.new_quantity_in_stock == 500 - analysis.sim_cards
    assert active_sim_row.new_quantity_dispatched == 100 + analysis.sim_cards
    assert historical_sim_row.changes_required is False
    assert next_sim_row.changes_required is False


def test_summary_analysis_builds_preview_metadata_and_detects_mismatch(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    create_daily_sheet(
        wb,
        'DAILY OUTPUT FILE 03-07-2026',
        [
            [1, datetime(2026, 7, 3).date(), 'BR003', 'Branch C', 'SIM', 2, 'WB003', '30001|30002', "'00034567"],
        ]
    )
    create_summary_sheet(
        wb,
        [
            ['SIM Orders', 1, None, None, 'Existing value differs'],
            ['SIM Cards', 400, None, None, 'Existing matches calculated cards'],
        ]
    )
    path = tmp_path / 'preview_metadata.xlsx'
    wb.save(path)

    engine = SummaryReconciliationEngine()
    analysis = engine.analyze(str(path))

    assert analysis.preview_ready is True
    assert analysis.has_changes is True
    assert analysis.changes_count == 1
    assert analysis.current_summary_values['SIM Orders'] == 1
    assert analysis.calculated_summary_values['SIM Orders'] == 2
    mismatch = next(diff for diff in analysis.differences if diff.metric_name == 'SIM Orders')
    assert mismatch.status == 'Mismatch'
    assert mismatch.existing_value == 1
    assert mismatch.calculated_value == 2
    assert mismatch.difference == 1


def test_summary_analysis_reports_match_when_summary_values_align(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    create_daily_sheet(
        wb,
        'DAILY OUTPUT FILE 03-07-2026',
        [
            [1, datetime(2026, 7, 3).date(), 'BR003', 'Branch C', 'SIM', 2, 'WB003', '30001|30002', "'00034567"],
        ]
    )
    create_summary_sheet(
        wb,
        [
            ['SIM Orders', 2, None, None, 'Matches calculated summary'],
            ['SIM Cards', 400, None, None, 'Matches on summary cards']
        ]
    )
    path = tmp_path / 'preview_match.xlsx'
    wb.save(path)

    engine = SummaryReconciliationEngine()
    analysis = engine.analyze(str(path))

    assert analysis.preview_ready is True
    assert analysis.has_changes is False
    assert analysis.changes_count == 0
    match = next(diff for diff in analysis.differences if diff.metric_name == 'SIM Orders')
    assert match.status == 'Match'
    assert match.difference == 0
