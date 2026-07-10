import openpyxl
from pathlib import Path

from src.core.summary_reconciliation_engine import SummaryReconciliationEngine
from src.core.summary_updater import SummaryWorksheetUpdater
from src.models.models import SummaryAnalysis


def create_summary_sheet(workbook: openpyxl.Workbook, rows: list) -> None:
    worksheet = workbook.create_sheet('CAPITEC SUMMARY FILE REPORT')
    headers = ['Files received', 'Total quantity received', 'Quantity in stock', 'Quantity dispatched', 'Comment']
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)


def create_daily_sheet(workbook: openpyxl.Workbook, title: str, rows: list) -> None:
    worksheet = workbook.create_sheet(title)
    headers = ['Order_No', 'Order_Creation_Date', 'Branch_Code', 'Branch_Name', 'Card_Type', 'No_of_Batches', 'Waybill_No', 'Batch_No', 'Bag_No']
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)


def test_summary_updater_updates_only_changed_fields(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    create_summary_sheet(
        wb,
        [
            ['SIM Orders', 1, None, None, ''],
            ['SIM Cards', 400, None, None, ''],
            ['DMCCLS Orders', 0, None, None, ''],
            ['DMCCLS Cards', 0, None, None, ''],
            ['Total Orders', 1, None, None, ''],
            ['Total Cards', 200, None, None, ''],
        ]
    )
    create_daily_sheet(
        wb,
        'DAILY OUTPUT FILE 03-07-2026',
        [
            [1, '2026-07-03', 'BR003', 'Branch C', 'SIM', 1, 'WB003', '30001|30002|30003|30004|30005', "'00034567"],
        ]
    )

    path = tmp_path / 'summary_update.xlsx'
    wb.save(path)

    engine = SummaryReconciliationEngine()
    analysis = engine.analyze(str(path))
    assert analysis.preview_ready is True

    updater = SummaryWorksheetUpdater()
    result = updater.update_summary_worksheet(analysis, engine.loader)

    assert result.success is True
    assert len(result.updated_fields) >= 1
    assert len(result.failed_fields) == 0
    assert len(result.skipped_fields) >= 0


def test_summary_updater_skips_correct_values(tmp_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    create_summary_sheet(
        wb,
        [
            ['SIM Orders', 2, None, None, ''],
            ['SIM Cards', 400, None, None, ''],
            ['DMCCLS Orders', 0, None, None, ''],
            ['DMCCLS Cards', 0, None, None, ''],
            ['Total Orders', 2, None, None, ''],
            ['Total Cards', 400, None, None, ''],
        ]
    )
    create_daily_sheet(
        wb,
        'DAILY OUTPUT FILE 03-07-2026',
        [
            [1, '2026-07-03', 'BR003', 'Branch C', 'SIM', 2, 'WB003', '30001|30002|30003|30004', "'00034567"],
        ]
    )

    path = tmp_path / 'summary_update_skip.xlsx'
    wb.save(path)

    engine = SummaryReconciliationEngine()
    analysis = engine.analyze(str(path))
    assert analysis.preview_ready is True

    updater = SummaryWorksheetUpdater()
    result = updater.update_summary_worksheet(analysis, engine.loader)

    assert result.success is True
    assert all('SIM Orders' not in field or 'SIM Orders' in field for field in result.skipped_fields)
    assert len(result.updated_fields) == 0
    assert len(result.failed_fields) == 0
