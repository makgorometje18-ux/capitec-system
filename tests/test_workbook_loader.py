import openpyxl
from pathlib import Path
from src.core.workbook_loader import WorkbookLoader


def test_single_daily_output_sheet(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DAILY OUTPUT FILE 01-07-2026"
    path = tmp_path / "single_daily.xlsx"
    wb.save(path)

    loader = WorkbookLoader()
    loader.load_workbook(str(path))
    selected = loader.detect_daily_output_sheet()

    assert selected == "DAILY OUTPUT FILE 01-07-2026"


def test_multiple_daily_output_sheets_selects_newest(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DAILY OUTPUT FILE 01-07-2026"
    wb.create_sheet("DAILY OUTPUT FILE 03-07-2026")
    wb.create_sheet("DAILY OUTPUT FILE 02-07-2026")
    path = tmp_path / "multiple_daily.xlsx"
    wb.save(path)

    loader = WorkbookLoader()
    loader.load_workbook(str(path))
    selected = loader.detect_daily_output_sheet()

    assert selected == "DAILY OUTPUT FILE 03-07-2026"


def test_daily_output_sheet_with_trailing_spaces(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DAILY OUTPUT FILE 01-07-2026 "
    path = tmp_path / "trailing_spaces.xlsx"
    wb.save(path)

    loader = WorkbookLoader()
    loader.load_workbook(str(path))
    selected = loader.detect_daily_output_sheet()

    assert selected == "DAILY OUTPUT FILE 01-07-2026 "


def test_invalid_daily_output_sheet_names_are_skipped(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DAILY OUTPUT FILE INVALID"
    wb.create_sheet("DAILY OUTPUT FILE 05-07-2026")
    path = tmp_path / "invalid_names.xlsx"
    wb.save(path)

    loader = WorkbookLoader()
    loader.load_workbook(str(path))
    selected = loader.detect_daily_output_sheet()

    assert selected == "DAILY OUTPUT FILE 05-07-2026"


def test_mixed_worksheet_names(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SAMPLE SHEET"
    wb.create_sheet("DAILY OUTPUT FILE 01-07-2026")
    wb.create_sheet("NOT DAILY OUTPUT FILE 02-07-2026")
    wb.create_sheet("DAILY OUTPUT FILE 03-07-2026")
    path = tmp_path / "mixed_names.xlsx"
    wb.save(path)

    loader = WorkbookLoader()
    loader.load_workbook(str(path))
    selected = loader.detect_daily_output_sheet()

    assert selected == "DAILY OUTPUT FILE 03-07-2026"
